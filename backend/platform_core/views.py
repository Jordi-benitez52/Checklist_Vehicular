from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db import models
from django.db.models import Q, Count
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from rest_framework import generics, status
from rest_framework.permissions import IsAuthenticated, BasePermission
from rest_framework.response import Response
from rest_framework.views import APIView


class IsAdminRole(BasePermission):
    """
    Permiso personalizado: Solo usuarios con role='admin' pueden acceder.
    """

    def has_permission(self, request, view):
        if not request.user or not request.user.is_authenticated:
            return False
        try:
            profile = request.user.profile
            return profile.role == 'admin'
        except Exception:
            return False

import json

from .models import (
    Turno,
    Empleado,
    Conductor,
    Vehiculo,
    AsignacionVehiculoEmpleado,
    AsignacionConductorVehiculo,
    AsignacionEmpleadoVehiculo,
    HistorialUsoVehiculo,
    BitacoraCambios,
    VisitanteRegistro,
    PlantillaChecklist,
    RegistroAcceso,
    ChecklistRegistro,
    ChecklistItemResultado,
    ChecklistTracto,
    ChecklistTractoItemCatalogo,
    ChecklistTractoResultado,
    ChecklistTractoLlanta,
    ChecklistTractoEvidencia,
    AuditLog,
    Notificacion,
)

from .serializers import (
    TurnoSerializer,
    EmpleadoSerializer,
    ConductorSerializer,
    VehiculoSerializer,
    AsignacionVehiculoEmpleadoSerializer,
    AsignacionConductorVehiculoSerializer,
    AsignacionEmpleadoVehiculoSerializer,
    HistorialUsoVehiculoSerializer,
    VisitanteRegistroSerializer,
    PlantillaChecklistSerializer,
    RegistroAccesoSerializer,
    ChecklistRegistroSerializer,
    ChecklistTractoSerializer,
    ChecklistTractoCreateSerializer,
    ChecklistTractoItemCatalogoSerializer,
    AuditLogSerializer,
    BitacoraCambiosSerializer,
    NotificacionSerializer,
    NotificacionCreateSerializer,
)

User = get_user_model()


# =========================
# HELPERS
# =========================

def registrar_auditoria(
    usuario,
    modulo,
    accion,
    descripcion,
    entidad_tipo='',
    entidad_id=None,
    turno=None
):
    try:
        AuditLog.objects.create(
            usuario=usuario if usuario and usuario.is_authenticated else None,
            modulo=modulo,
            accion=accion,
            descripcion=descripcion,
            entidad_tipo=entidad_tipo,
            entidad_id=entidad_id,
            turno=turno
        )
    except Exception:
        pass


def parse_datetime_form(value):
    if not value:
        return timezone.now()

    dt = parse_datetime(value)

    if dt is None:
        return timezone.now()

    if timezone.is_naive(dt):
        dt = timezone.make_aware(dt)

    return dt


def obtener_perfil_usuario(usuario):
    if hasattr(usuario, 'profile'):
        return usuario.profile

    if hasattr(usuario, 'platform_profile'):
        return usuario.platform_profile

    return None


def es_admin(user):
    try:
        if user.is_superuser:
            return True
        perfil = obtener_perfil_usuario(user)
        if perfil is None:
            return False
        return perfil.role == 'admin'
    except Exception:
        return False


# =========================
# API VIEWS - CATÁLOGOS
# =========================

class VehiculoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VehiculoSerializer

    def get_queryset(self):
        queryset = Vehiculo.objects.filter(activo=True).order_by('placa')

        tipo_entidad = self.request.query_params.get('tipo_entidad')
        categoria = self.request.query_params.get('categoria')
        placa = self.request.query_params.get('placa')

        if tipo_entidad:
            queryset = queryset.filter(tipo_entidad=tipo_entidad)

        if categoria:
            queryset = queryset.filter(categoria=categoria)

        if placa:
            queryset = queryset.filter(placa__icontains=placa)

        return queryset


class VehiculoCreateAPIView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VehiculoSerializer


class VehiculoDetailAPIView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VehiculoSerializer
    queryset = Vehiculo.objects.all()


class VehiculoUpdateAPIView(generics.UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VehiculoSerializer
    queryset = Vehiculo.objects.all()


class VehiculoDesactivarAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        vehiculo = get_object_or_404(Vehiculo, pk=pk)
        vehiculo.activo = False
        vehiculo.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='editar',
            descripcion=f'Se desactivó vehículo {vehiculo.placa}.',
            entidad_tipo='Vehiculo',
            entidad_id=vehiculo.id,
            turno=None
        )

        return Response(
            {'message': 'Vehículo desactivado correctamente.'},
            status=status.HTTP_200_OK
        )


class EmpleadoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EmpleadoSerializer
    queryset = Empleado.objects.filter(activo=True).order_by('nombre_completo')


class ConductorListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ConductorSerializer
    queryset = Conductor.objects.filter(activo=True).order_by('nombre_completo')


class ConductorCreateAPIView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ConductorSerializer

    def create(self, request, *args, **kwargs):
        data = request.data
        nombre_completo = data.get('nombre_completo', '').strip()
        licencia = data.get('licencia', '').strip().upper()
        telefono = data.get('telefono', '').strip() or None
        empresa = data.get('empresa', '').strip() or None

        if not nombre_completo:
            return Response(
                {'nombre_completo': ['El nombre completo es requerido']},
                status=status.HTTP_400_BAD_REQUEST
            )

        if licencia and Conductor.objects.filter(licencia=licencia, activo=True).exists():
            return Response(
                {'licencia': ['Ya existe un conductor con esta licencia']},
                status=status.HTTP_400_BAD_REQUEST
            )

        conductor = Conductor.objects.create(
            nombre_completo=nombre_completo,
            licencia=licencia or None,
            telefono=telefono,
            empresa=empresa
        )

        BitacoraCambios.objects.create(
            usuario=request.user,
            accion='CREATE',
            modulo='CONDUCTORES',
            descripcion=f'Se creó conductor {conductor.nombre_completo} (licencia: {conductor.licencia})',
            detalle=f'ID: {conductor.id}'
        )

        serializer = self.get_serializer(conductor)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class EmpleadoCreateAPIView(generics.CreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EmpleadoSerializer

    def create(self, request, *args, **kwargs):
        data = request.data
        numero_empleado = data.get('numero_empleado', '').strip().upper()
        nombre_completo = data.get('nombre_completo', '').strip()
        departamento = data.get('departamento', '').strip() or None
        puesto = data.get('puesto', '').strip() or None

        if not numero_empleado:
            return Response(
                {'numero_empleado': ['El número de empleado es requerido']},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not nombre_completo:
            return Response(
                {'nombre_completo': ['El nombre completo es requerido']},
                status=status.HTTP_400_BAD_REQUEST
            )

        if Empleado.objects.filter(numero_empleado=numero_empleado, activo=True).exists():
            return Response(
                {'numero_empleado': ['Ya existe un empleado con este número']},
                status=status.HTTP_400_BAD_REQUEST
            )

        empleado = Empleado.objects.create(
            numero_empleado=numero_empleado,
            nombre_completo=nombre_completo,
            departamento=departamento,
            puesto=puesto
        )

        BitacoraCambios.objects.create(
            usuario=request.user,
            accion='CREATE',
            modulo='EMPLEADOS',
            descripcion=f'Se creó empleado {empleado.nombre_completo} ({empleado.numero_empleado})',
            detalle=f'ID: {empleado.id}'
        )

        serializer = self.get_serializer(empleado)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ConductorUpdateAPIView(generics.UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ConductorSerializer
    queryset = Conductor.objects.all()

    def update(self, request, *args, **kwargs):
        conductor = self.get_object()
        data = request.data

        nombre_completo = data.get('nombre_completo', '').strip()
        licencia = data.get('licencia', '').strip().upper() or None
        telefono = data.get('telefono', '').strip() or None
        empresa = data.get('empresa', '').strip() or None

        if nombre_completo:
            conductor.nombre_completo = nombre_completo
        if 'licencia' in data:
            conductor.licencia = licencia
        if 'telefono' in data:
            conductor.telefono = telefono
        if 'empresa' in data:
            conductor.empresa = empresa

        conductor.save()

        BitacoraCambios.objects.create(
            usuario=request.user,
            accion='UPDATE',
            modulo='CONDUCTORES',
            descripcion=f'Se actualizó conductor {conductor.nombre_completo}',
            detalle=f'ID: {conductor.id}'
        )

        serializer = self.get_serializer(conductor)
        return Response(serializer.data)


class ConductorDesactivarAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        conductor = get_object_or_404(Conductor, pk=pk)
        conductor.activo = False
        conductor.save()

        BitacoraCambios.objects.create(
            usuario=request.user,
            accion='DEACTIVATE',
            modulo='CONDUCTORES',
            descripcion=f'Se desactivó conductor {conductor.nombre_completo}',
            detalle=f'ID: {conductor.id}'
        )

        return Response({'message': 'Conductor desactivado correctamente'})


class EmpleadoUpdateAPIView(generics.UpdateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = EmpleadoSerializer
    queryset = Empleado.objects.all()

    def update(self, request, *args, **kwargs):
        empleado = self.get_object()
        data = request.data

        numero_empleado = data.get('numero_empleado', '').strip().upper()
        nombre_completo = data.get('nombre_completo', '').strip()
        departamento = data.get('departamento', '').strip() or None
        puesto = data.get('puesto', '').strip() or None

        if numero_empleado:
            empleado.numero_empleado = numero_empleado
        if nombre_completo:
            empleado.nombre_completo = nombre_completo
        if 'departamento' in data:
            empleado.departamento = departamento
        if 'puesto' in data:
            empleado.puesto = puesto

        empleado.save()

        BitacoraCambios.objects.create(
            usuario=request.user,
            accion='UPDATE',
            modulo='EMPLEADOS',
            descripcion=f'Se actualizó empleado {empleado.nombre_completo}',
            detalle=f'ID: {empleado.id}'
        )

        serializer = self.get_serializer(empleado)
        return Response(serializer.data)


class EmpleadoDesactivarAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        empleado_id = request.query_params.get('id')
        if not empleado_id:
            return Response({'error': 'ID requerido'}, status=status.HTTP_400_BAD_REQUEST)

        empleado = get_object_or_404(Empleado, pk=empleado_id)
        empleado.activo = False
        empleado.save()

        BitacoraCambios.objects.create(
            usuario=request.user,
            accion='DEACTIVATE',
            modulo='EMPLEADOS',
            descripcion=f'Se desactivó empleado {empleado.nombre_completo}',
            detalle=f'ID: {empleado.id}'
        )

        return Response({'message': 'Empleado desactivado correctamente'})


class AsignacionVehiculoEmpleadoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AsignacionVehiculoEmpleadoSerializer
    queryset = AsignacionVehiculoEmpleado.objects.filter(activa=True).select_related(
        'vehiculo',
        'empleado'
    ).order_by('-fecha_asignacion')


class EmpleadosConVehiculoAsignadoAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        asignaciones = AsignacionVehiculoEmpleado.objects.filter(
            activa=True
        ).select_related('vehiculo', 'empleado').order_by('empleado__nombre_completo')

        data = []
        for a in asignaciones:
            data.append({
                'id': a.empleado.id,
                'numero_empleado': a.empleado.numero_empleado,
                'nombre_completo': a.empleado.nombre_completo,
                'departamento': a.empleado.departamento,
                'puesto': a.empleado.puesto,
                'vehiculo_id': a.vehiculo.id,
                'vehiculo_placa': a.vehiculo.placa,
                'vehiculo_clave_interna': a.vehiculo.clave_interna,
                'vehiculo_marca': a.vehiculo.marca,
                'vehiculo_modelo': a.vehiculo.modelo,
                'asignacion_id': a.id,
            })

        return Response(data)


class VisitanteRegistroListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = VisitanteRegistroSerializer
    queryset = VisitanteRegistro.objects.all().order_by('-id')


# =========================
# API VIEWS - TURNOS
# =========================

class TurnoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = TurnoSerializer

    def get_queryset(self):
        queryset = Turno.objects.select_related('guardia').all().order_by('-fecha', '-hora_apertura')

        abierto = self.request.query_params.get('abierto')
        guardia = self.request.query_params.get('guardia')

        if abierto is not None:
            if abierto.lower() == 'true':
                queryset = queryset.filter(abierto=True)
            elif abierto.lower() == 'false':
                queryset = queryset.filter(abierto=False)

        if guardia:
            queryset = queryset.filter(guardia_id=guardia)

        return queryset


class TurnoCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import sys

        print(f"[DEBUG] === TurnoCreateAPIView called ===", file=sys.stderr)
        print(f"[DEBUG] request.data keys: {list(request.data.keys())}", file=sys.stderr)
        for key in request.data.keys():
            val = request.data.get(key)
            if hasattr(val, 'size'):
                print(f"  {key}: FILE(name={val.name}, size={val.size})", file=sys.stderr)
            elif isinstance(val, str) and len(val) > 80:
                print(f"  {key}: str(len={len(val)}, preview={val[:50]}...)", file=sys.stderr)
            else:
                print(f"  {key}: {type(val).__name__} = {repr(val)[:100]}", file=sys.stderr)

        raw_data = request.data
        if hasattr(raw_data, 'lists'):
            data = {k: raw_data.getlist(k)[0] if len(raw_data.getlist(k)) == 1 else raw_data.getlist(k) for k in raw_data.keys()}
        elif hasattr(raw_data, 'dict'):
            data = raw_data.dict()
        else:
            data = dict(raw_data)

        is_admin = hasattr(request.user, 'profile') and request.user.profile.role == 'admin'

        if is_admin and data.get('guardia_id'):
            from django.contrib.auth.models import User
            try:
                guardia_obj = User.objects.get(pk=data.get('guardia_id'))
            except User.DoesNotExist:
                return Response({'error': 'El guardia seleccionado no existe.'}, status=status.HTTP_400_BAD_REQUEST)
        else:
            guardia_obj = request.user

        abierta_existente = Turno.objects.filter(
            abierto=True
        ).first()

        if abierta_existente:
            return Response(
                {
                    'error': 'Ya existe un turno abierto. Debes cerrarlo antes de abrir otro.',
                    'turno_actual': {
                        'id': abierta_existente.id,
                        'tipo_turno': abierta_existente.tipo_turno,
                        'fecha': abierta_existente.fecha,
                        'guardia': abierta_existente.guardia.username
                    }
                },
                status=status.HTTP_400_BAD_REQUEST
            )

        turno_data = {
            'tipo_turno': data.get('tipo_turno'),
            'fecha': data.get('fecha'),
            'abierto': True,
            'observaciones': data.get('observaciones', ''),
        }

        if not turno_data['tipo_turno']:
            return Response({'error': 'El tipo de turno es requerido.'}, status=status.HTTP_400_BAD_REQUEST)

        if not turno_data['fecha']:
            return Response({'error': 'La fecha es requerida.'}, status=status.HTTP_400_BAD_REQUEST)

        turno_data['hora_apertura'] = timezone.now().isoformat()

        Turno.objects.filter(abierto=True).update(abierto=False)

        serializer = TurnoSerializer(data=turno_data)

        if serializer.is_valid():
            from django.utils import timezone as tz
            turno = Turno.objects.create(
                guardia=guardia_obj,
                tipo_turno=turno_data['tipo_turno'],
                fecha=turno_data['fecha'],
                hora_apertura=tz.now(),
                abierto=True,
                observaciones=turno_data.get('observaciones', '')
            )

            registrar_auditoria(
                usuario=request.user,
                modulo='turnos',
                accion='crear',
                descripcion=f'Se abrió turno {turno.id} de tipo {turno.tipo_turno} para {guardia_obj.username}.',
                entidad_tipo='Turno',
                entidad_id=turno.id,
                turno=turno
            )

            try:
                from django.contrib.auth.models import User
                all_users = User.objects.filter(is_active=True, profile__is_active_user=True)
                for user in all_users:
                    Notificacion.objects.create(
                        usuario=user,
                        titulo='Turno abierto',
                        mensaje=f'Se ha abierto el turno {turno.get_tipo_turno_display()} ({turno.fecha}) por {request.user.username}',
                        tipo='turno_abierto',
                        link=f'/turnos/{turno.id}'
                    )
            except Exception:
                pass

            return Response(
                {
                    'message': 'Turno creado correctamente.',
                    'data': TurnoSerializer(turno).data
                },
                status=status.HTTP_201_CREATED
            )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class TurnoCloseAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            turno = Turno.objects.get(pk=pk, abierto=True)
        except Turno.DoesNotExist:
            return Response(
                {'error': 'Turno no encontrado o ya está cerrado.'},
                status=status.HTTP_404_NOT_FOUND
            )

        turno.abierto = False
        turno.hora_cierre = timezone.now()

        observaciones = request.data.get('observaciones_cierre', '')
        if observaciones:
            turno.observaciones_cierre = observaciones

        firma_data = request.data.get('firma_cierre', '')
        if firma_data:
            turno.firma_cierre = firma_data

        turno.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='turnos',
            accion='cerrar',
            descripcion=f'Se cerró el turno {turno.id}.',
            entidad_tipo='Turno',
            entidad_id=turno.id,
            turno=turno
        )

        try:
            from django.contrib.auth.models import User
            all_users = User.objects.filter(is_active=True, profile__is_active_user=True)
            for user in all_users:
                Notificacion.objects.create(
                    usuario=user,
                    titulo='Turno cerrado',
                    mensaje=f'Se ha cerrado el turno {turno.get_tipo_turno_display()} ({turno.fecha}) por {request.user.username}',
                    tipo='turno_cerrado',
                    link=f'/turnos/{turno.id}'
                )
        except Exception:
            pass

        return Response({
            'message': 'Turno cerrado correctamente.',
            'data': TurnoSerializer(turno).data
        })


# =========================
# API VIEWS - PLANTILLAS / CHECKLIST GENERAL
# =========================

class PlantillaChecklistListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = PlantillaChecklistSerializer

    def get_queryset(self):
        queryset = PlantillaChecklist.objects.filter(
            activa=True
        ).prefetch_related('items').order_by('nombre')

        tipo_entidad = self.request.query_params.get('tipo_entidad')
        categoria = self.request.query_params.get('categoria')

        if tipo_entidad:
            queryset = queryset.filter(tipo_entidad=tipo_entidad)

        if categoria:
            queryset = queryset.filter(categoria=categoria)

        return queryset


class ChecklistRegistroListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ChecklistRegistroSerializer

    def get_queryset(self):
        queryset = ChecklistRegistro.objects.select_related(
            'registro_acceso',
            'plantilla',
            'registrado_por',
        ).prefetch_related('resultados').order_by('-fecha_hora')

        plantilla = self.request.query_params.get('plantilla')
        resultado = self.request.query_params.get('resultado')

        if plantilla:
            queryset = queryset.filter(plantilla_id=plantilla)

        if resultado:
            queryset = queryset.filter(resultado_general=resultado)

        return queryset


class ChecklistRegistroCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

def post(self, request):
        import sys
        print(f"[DEBUG] === ChecklistRegistroCreateAPIView called ===", file=sys.stderr)
        print(f"[DEBUG] request.data keys: {list(request.data.keys())}", file=sys.stderr)

        resultados = request.data.get('resultados', [])
        raw_data = request.data
        if hasattr(raw_data, 'lists'):
            data = {k: raw_data.getlist(k)[0] if len(raw_data.getlist(k)) == 1 else raw_data.getlist(k) for k in raw_data.keys()}
        elif hasattr(raw_data, 'dict'):
            data = raw_data.dict()
        else:
            data = dict(raw_data)

        serializer = ChecklistRegistroSerializer(data=data)

        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        checklist = serializer.save(registrado_por=request.user)

        for resultado in resultados:
            ChecklistItemResultado.objects.create(
                checklist=checklist,
                item_plantilla_id=resultado.get('item_plantilla'),
                valor_booleano=resultado.get('valor_booleano'),
                valor_texto=resultado.get('valor_texto'),
                observacion=resultado.get('observacion'),
            )

        registrar_auditoria(
            usuario=request.user,
            modulo='accesos',
            accion='crear',
            descripcion=f'Se creó checklist general {checklist.id}.',
            entidad_tipo='ChecklistRegistro',
            entidad_id=checklist.id,
            turno=checklist.registro_acceso.turno if checklist.registro_acceso else None
        )

        checklist_refresh = ChecklistRegistro.objects.select_related(
            'registro_acceso',
            'plantilla',
            'registrado_por',
        ).prefetch_related('resultados').get(id=checklist.id)

        return Response(
            {
                'message': 'Checklist registrado correctamente.',
                'data': ChecklistRegistroSerializer(
                    checklist_refresh,
                    context={'request': request}
                ).data
            },
            status=status.HTTP_201_CREATED
        )


# =========================
# API VIEWS - REGISTRO DE ACCESO
# =========================

class RegistroAccesoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = RegistroAccesoSerializer

    def get_queryset(self):
        queryset = RegistroAcceso.objects.select_related(
            'turno',
            'guardia',
            'vehiculo',
            'empleado',
            'conductor',
            'visitante',
        ).order_by('-fecha_hora')

        tipo_movimiento = self.request.query_params.get('tipo_movimiento')
        tipo_entidad = self.request.query_params.get('tipo_entidad')
        vehiculo = self.request.query_params.get('vehiculo')
        visitante = self.request.query_params.get('visitante')
        turno = self.request.query_params.get('turno')
        checklist_pendiente = self.request.query_params.get('checklist_pendiente')

        if tipo_movimiento:
            queryset = queryset.filter(tipo_movimiento=tipo_movimiento)

        if tipo_entidad:
            queryset = queryset.filter(tipo_entidad=tipo_entidad)

        if vehiculo:
            queryset = queryset.filter(vehiculo_id=vehiculo)

        if visitante:
            queryset = queryset.filter(visitante_id=visitante)

        if turno:
            queryset = queryset.filter(turno_id=turno)

        if checklist_pendiente and checklist_pendiente.lower() == 'true':
            queryset = queryset.filter(
                checklist_requerido=True,
                checklist_realizado=False
            )

        return queryset


class RegistroAccesoCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import sys

        print(f"[DEBUG] === RegistroAccesoCreateAPIView called ===", file=sys.stderr)
        print(f"[DEBUG] request.data keys: {list(request.data.keys())}", file=sys.stderr)
        for key in request.data.keys():
            val = request.data.get(key)
            if hasattr(val, 'size'):
                print(f"  {key}: FILE(name={val.name}, size={val.size})", file=sys.stderr)
            elif isinstance(val, str) and len(val) > 80:
                print(f"  {key}: str(len={len(val)}, preview={val[:50]}...)", file=sys.stderr)
            else:
                print(f"  {key}: {type(val).__name__} = {repr(val)[:100]}", file=sys.stderr)

        raw_data = request.data
        if hasattr(raw_data, 'lists'):
            data = {k: raw_data.getlist(k)[0] if len(raw_data.getlist(k)) == 1 else raw_data.getlist(k) for k in raw_data.keys()}
        elif hasattr(raw_data, 'dict'):
            data = raw_data.dict()
        else:
            data = dict(raw_data)

        tipo_movimiento = data.get('tipo_movimiento')
        vehiculo_id = data.get('vehiculo')
        turno_id = data.get('turno')
        tipo_entidad = data.get('tipo_entidad')
        conductor_id = data.get('conductor')
        empleado_id = data.get('empleado')
        visitante_id = data.get('visitante')

        from .models import RegistroAcceso, Vehiculo, Conductor, Empleado

        if tipo_movimiento == 'entrada':
            if tipo_entidad == 'tracto' and vehiculo_id and conductor_id:
                vehiculo = Vehiculo.objects.get(id=vehiculo_id)
                if vehiculo.en_instalacion:
                    return Response(
                        {'vehiculo': 'El tracto ya está dentro de la instalación.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # NUEVA VALIDACIÓN: Verificar usando tabla de asignaciones
                from platform_core.models import AsignacionConductorVehiculo
                asignacion_valida = AsignacionConductorVehiculo.objects.filter(
                    conductor_id=conductor_id,
                    vehiculo_id=vehiculo_id,
                    activa=True,
                    fecha_desasignacion__isnull=True
                ).exists()

                if not asignacion_valida:
                    return Response(
                        {'conductor': 'Este conductor no está asignado a este tractocamión. Contacte al administrador.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                conductor_ya_dentro = RegistroAcceso.objects.filter(
                    conductor_id=conductor_id,
                    tipo_movimiento='entrada',
                    conductor_pendiente_salida=True
                ).exists()
                if conductor_ya_dentro:
                    return Response(
                        {'conductor': f'El conductor ya tiene un registro de entrada activo.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                data['conductor_pendiente_salida'] = True

            elif tipo_entidad == 'empleado_propio':
                emp_nombre = data.get('empleado_propio_nombre', '').strip()
                emp_apellido = data.get('empleado_propio_apellido', '').strip()
                emp_placas = data.get('empleado_propio_placas', '').strip().upper()
                emp_marca = data.get('empleado_propio_marca', '').strip()

                if not emp_nombre or not emp_apellido or not emp_placas:
                    return Response(
                        {'empleado_propio': 'Nombre, apellido y placas son requeridos para registro de empleado con vehículo propio.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if len(emp_nombre) > 50:
                    return Response(
                        {'empleado_propio_nombre': 'El nombre debe tener máximo 50 caracteres.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if len(emp_apellido) > 50:
                    return Response(
                        {'empleado_propio_apellido': 'El apellido debe tener máximo 50 caracteres.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if len(emp_placas) > 10:
                    return Response(
                        {'empleado_propio_placas': 'Las placas deben tener máximo 10 caracteres.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if len(emp_marca) > 50:
                    return Response(
                        {'empleado_propio_marca': 'La marca/modelo debe tener máximo 50 caracteres.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                duplicado = RegistroAcceso.objects.filter(
                    tipo_movimiento='entrada',
                    tipo_entidad='empleado_propio'
                ).exclude(
                    id__in=RegistroAcceso.objects.filter(
                        tipo_movimiento='salida',
                        entrada_asociada__isnull=False
                    ).values_list('entrada_asociada_id', flat=True)
                )

                for reg in duplicado:
                    obs = reg.observaciones or ''
                    if '[EMPLEADO PROPIO]' in obs:
                        parts = obs.split('|')
                        if len(parts) >= 2:
                            nombre_part = parts[0].replace('[EMPLEADO PROPIO]', '').strip()
                            placas_en_obs = ''
                            for p in parts[1:]:
                                if 'Placas:' in p:
                                    placas_en_obs = p.split('Placas:')[1].strip().split('|')[0].strip()
                                    break

                            nombre_completo = f'{emp_nombre} {emp_apellido}'.upper()
                            if nombre_part.upper() == nombre_completo and placas_en_obs.upper() == emp_placas:
                                return Response(
                                    {'empleado_propio': f'Ya existe un registro de entrada activo para {emp_nombre} {emp_apellido} con placas {emp_placas}. Debe registrar la salida primero.'},
                                    status=status.HTTP_400_BAD_REQUEST
                                )

                data['conductor_pendiente_salida'] = True

            elif tipo_entidad == 'empleado' and empleado_id:
                vehiculo_id_int = int(vehiculo_id) if vehiculo_id else None

                if vehiculo_id_int:
                    vehiculo = Vehiculo.objects.get(id=vehiculo_id_int)
                    if vehiculo.en_instalacion:
                        return Response(
                            {'error': 'El empleado ya tiene un registro de entrada activo con este vehículo.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )
                else:
                    vehiculo = None

                if vehiculo:
                    data['vehiculo'] = vehiculo_id
                    data['conductor_pendiente_salida'] = False
                else:
                    data['conductor_pendiente_salida'] = True

            elif tipo_entidad == 'conductor':
                if not conductor_id:
                    return Response(
                        {'conductor': 'Debe seleccionar un conductor de la lista.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                conductor = Conductor.objects.get(id=conductor_id)
                if not conductor.activo:
                    return Response(
                        {'conductor': 'El conductor seleccionado no está activo.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                tiene_entrada_pendiente = RegistroAcceso.objects.filter(
                    conductor=conductor,
                    tipo_movimiento='entrada',
                    conductor_pendiente_salida=True
                ).exclude(
                    id__in=RegistroAcceso.objects.filter(
                        tipo_movimiento='salida',
                        entrada_asociada__isnull=False
                    ).values_list('entrada_asociada_id', flat=True)
                ).exists()

                if tiene_entrada_pendiente:
                    return Response(
                        {'conductor': f'El conductor {conductor.nombre_completo} ya tiene un registro de entrada activo.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                data['conductor_id'] = conductor_id
                data['conductor_pendiente_salida'] = True

                if conductor.vehiculo:
                    conductor.vehiculo.en_instalacion = True
                    conductor.vehiculo.save(update_fields=['en_instalacion'])

            elif tipo_entidad == 'visitante':
                visitante_nombre = data.get('visitante_nombre', '').strip()
                visitante_apellido = data.get('visitante_apellido', '').strip()
                visitante_placas = data.get('visitante_placas', '').strip().upper()

                if visitante_nombre and visitante_apellido and visitante_placas:
                    duplicado = RegistroAcceso.objects.filter(
                        tipo_movimiento='entrada',
                        tipo_entidad='visitante'
                    ).exclude(
                        id__in=RegistroAcceso.objects.filter(
                            tipo_movimiento='salida',
                            entrada_asociada__isnull=False
                        ).values_list('entrada_asociada_id', flat=True)
                    )

                    for reg in duplicado:
                        obs = reg.observaciones or ''
                        nombre_en_obs = ''
                        placas_en_obs = ''

                        if '[VISITANTE]' in obs:
                            parts = obs.split('|')
                            if len(parts) >= 2:
                                nombre_part = parts[0].replace('[VISITANTE]', '').strip()
                                if 'Placas:' in parts[1]:
                                    placas_part = parts[1].split('Placas:')[1].strip() if 'Placas:' in parts[1] else ''
                                if len(parts) >= 3 and 'Placas:' in parts[2]:
                                    placas_part = parts[2].split('Placas:')[1].strip() if 'Placas:' in parts[2] else ''

                                nombre_en_obs = nombre_part.replace('VISITANTE', '').strip()
                                for p in parts[1:]:
                                    if 'Placas:' in p:
                                        placas_en_obs = p.split('Placas:')[1].strip().split('|')[0].split('Motivo:')[0].strip()
                                        break

                        if nombre_en_obs.upper() == f'{visitante_nombre} {visitante_apellido}'.upper() and placas_en_obs.upper() == visitante_placas:
                            return Response(
                                {'error': f'Ya existe un registro de entrada activo para {visitante_nombre} {visitante_apellido} con placas {visitante_placas}. Debe registrar la salida primero.'},
                                status=status.HTTP_400_BAD_REQUEST
                            )

                if visitante_id:
                    tiene_entrada_sin_salida = RegistroAcceso.objects.filter(
                        visitante_id=visitante_id,
                        tipo_movimiento='entrada',
                        tipo_entidad='visitante'
                    ).exclude(
                        id__in=RegistroAcceso.objects.filter(
                            tipo_movimiento='salida',
                            entrada_asociada__isnull=False
                        ).values_list('entrada_asociada_id', flat=True)
                    ).exists()
                    if tiene_entrada_sin_salida:
                        return Response(
                            {'visitante': 'Este visitante aún no ha registrado su salida.'},
                            status=status.HTTP_400_BAD_REQUEST
                        )

                data['conductor_pendiente_salida'] = True

        elif tipo_movimiento == 'salida':
            if tipo_entidad == 'tracto' and vehiculo_id and conductor_id:
                vehiculo = Vehiculo.objects.get(id=vehiculo_id)

                if not vehiculo.en_instalacion:
                    return Response(
                        {'vehiculo': 'Este tracto no está dentro de la instalación.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                if vehiculo.conductor_actual_id != int(conductor_id):
                    return Response(
                        {'conductor': 'El conductor no es el conductor actual de este tracto.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None
                if not evidencia_file:
                    return Response(
                        {'evidencia_fotografica': 'La evidencia fotográfica es obligatoria para la salida de tractocamiones.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            elif tipo_entidad == 'empleado' and vehiculo_id:
                vehiculo = Vehiculo.objects.get(id=vehiculo_id)
                if not vehiculo.en_instalacion:
                    return Response(
                        {'error': 'Este vehículo no está dentro de la instalación.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                entrada_pendiente_empleado = RegistroAcceso.objects.filter(
                    tipo_entidad='empleado',
                    tipo_movimiento='entrada',
                    vehiculo_id=vehiculo_id,
                    conductor_pendiente_salida=True
                ).exclude(
                    id__in=RegistroAcceso.objects.filter(
                        tipo_movimiento='salida',
                        entrada_asociada__isnull=False
                    ).values_list('entrada_asociada_id', flat=True)
                ).first()
                if entrada_pendiente_empleado:
                    data['entrada_asociada'] = entrada_pendiente_empleado.id

            elif tipo_entidad == 'conductor' and not conductor_id:
                entrada_asociada_id = data.get('entrada_asociada_id')
                entrada_pendiente = None

                if entrada_asociada_id:
                    try:
                        entrada_pendiente = RegistroAcceso.objects.get(
                            id=int(entrada_asociada_id),
                            tipo_entidad='conductor',
                            tipo_movimiento='entrada',
                            conductor_pendiente_salida=True
                        )
                    except (RegistroAcceso.DoesNotExist, ValueError):
                        entrada_pendiente = None

                if not entrada_pendiente:
                    conductor_nombre = data.get('conductor_nombre', '').strip()
                    conductor_apellido = data.get('conductor_apellido', '').strip()
                    nombre_completo = f'{conductor_nombre} {conductor_apellido}'.upper()

                    if nombre_completo.strip():
                        for reg in RegistroAcceso.objects.filter(
                            tipo_entidad='conductor',
                            tipo_movimiento='entrada',
                            conductor_pendiente_salida=True
                        ):
                            obs = reg.observaciones or ''
                            if '[CONDUCTOR]' in obs:
                                parts = obs.split('|')
                                if len(parts) >= 1:
                                    nombre_part = parts[0].replace('[CONDUCTOR]', '').strip().upper()
                                    if nombre_part == nombre_completo:
                                        entrada_pendiente = reg
                                        break

                if not entrada_pendiente:
                    return Response(
                        {'conductor': 'Este conductor no tiene un registro de entrada activo. Primero registre la entrada.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                data['entrada_asociada'] = entrada_pendiente.id

                evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None
                if not evidencia_file:
                    return Response(
                        {'evidencia_fotografica': 'La evidencia fotográfica es obligatoria para la salida de conductores.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                conductor_placa_salida = data.get('conductor_placa', '').strip().upper()
                conductor_marca_salida = data.get('conductor_marca', '').strip()

            elif tipo_entidad == 'conductor' and conductor_id:
                entrada_pendiente = RegistroAcceso.objects.filter(
                    conductor_id=conductor_id,
                    tipo_movimiento='entrada',
                    conductor_pendiente_salida=True
                ).first()
                if not entrada_pendiente:
                    return Response(
                        {'conductor': 'Este conductor no tiene un registro de entrada activo. Primero registre la entrada.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                data['entrada_asociada'] = entrada_pendiente.id
                evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None
                if not evidencia_file:
                    return Response(
                        {'evidencia_fotografica': 'La evidencia fotográfica es obligatoria para la salida de conductores.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                conductor_placa_salida = data.get('conductor_placa', '').strip().upper()
                conductor_marca_salida = data.get('conductor_marca', '').strip()

            elif tipo_entidad == 'visitante':
                evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None
                if not evidencia_file:
                    return Response(
                        {'evidencia_fotografica': 'La evidencia fotográfica es obligatoria para la salida de visitantes.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                entrada_asociada_id = data.get('entrada_asociada_id')
                entrada_pendiente = None

                if entrada_asociada_id:
                    try:
                        entrada_pendiente = RegistroAcceso.objects.get(
                            id=int(entrada_asociada_id),
                            tipo_entidad='visitante',
                            tipo_movimiento='entrada',
                            conductor_pendiente_salida=True
                        )
                    except (RegistroAcceso.DoesNotExist, ValueError):
                        entrada_pendiente = None

                if not entrada_pendiente and visitante_id:
                    entrada_pendiente = RegistroAcceso.objects.filter(
                        visitante_id=visitante_id,
                        tipo_entidad='visitante',
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).exclude(
                        id__in=RegistroAcceso.objects.filter(
                            tipo_movimiento='salida',
                            entrada_asociada__isnull=False
                        ).values_list('entrada_asociada_id', flat=True)
                    ).first()

                if not entrada_pendiente:
                    vis_nombre = data.get('visitante_nombre', '').strip().upper()
                    vis_apellido = data.get('visitante_apellido', '').strip().upper()
                    vis_placas = data.get('visitante_placas', '').strip().upper()
                    nombre_completo = f'{vis_nombre} {vis_apellido}'.strip()

                    if nombre_completo:
                        for reg in RegistroAcceso.objects.filter(
                            tipo_entidad='visitante',
                            tipo_movimiento='entrada',
                            conductor_pendiente_salida=True
                        ).exclude(
                            id__in=RegistroAcceso.objects.filter(
                                tipo_movimiento='salida',
                                entrada_asociada__isnull=False
                            ).values_list('entrada_asociada_id', flat=True)
                        ):
                            obs = reg.observaciones or ''
                            if '[VISITANTE]' in obs:
                                parts = obs.split('|')
                                if parts:
                                    nombre_part = parts[0].replace('[VISITANTE]', '').strip().upper()
                                    if nombre_part == nombre_completo:
                                        if vis_placas:
                                            for p in parts:
                                                if 'Placas:' in p:
                                                    placas_part = p.split('Placas:')[1].strip().split('|')[0].strip().upper()
                                                    if placas_part == vis_placas:
                                                        entrada_pendiente = reg
                                                        break
                                        else:
                                            entrada_pendiente = reg
                                        break

                if entrada_pendiente:
                    data['entrada_asociada'] = entrada_pendiente.id
                else:
                    return Response(
                        {'visitante': 'Este visitante no tiene un registro de entrada activo. Primero registre la entrada.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                visitante_placa_salida = data.get('visitante_placas', '').strip().upper()

            elif tipo_entidad == 'empleado_propio':
                evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None
                if not evidencia_file:
                    return Response(
                        {'evidencia_fotografica': 'La evidencia fotográfica es obligatoria para la salida de empleados con vehículo propio.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

                emp_nombre = data.get('empleado_propio_nombre', '').strip().upper()
                emp_apellido = data.get('empleado_propio_apellido', '').strip().upper()
                emp_placas = data.get('empleado_propio_placas', '').strip().upper()
                nombre_completo = f'{emp_nombre} {emp_apellido}'.strip()

                entrada_asociada_id = data.get('entrada_asociada_id')
                entrada_pendiente = None

                if entrada_asociada_id:
                    try:
                        entrada_pendiente = RegistroAcceso.objects.get(
                            id=int(entrada_asociada_id),
                            tipo_entidad='empleado',
                            tipo_movimiento='entrada',
                            conductor_pendiente_salida=True
                        )
                    except (RegistroAcceso.DoesNotExist, ValueError):
                        entrada_pendiente = None

                if not entrada_pendiente and nombre_completo:
                    for reg in RegistroAcceso.objects.filter(
                        tipo_entidad='empleado',
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).exclude(
                        id__in=RegistroAcceso.objects.filter(
                            tipo_movimiento='salida',
                            entrada_asociada__isnull=False
                        ).values_list('entrada_asociada_id', flat=True)
                    ):
                        obs = reg.observaciones or ''
                        if '[EMPLEADO PROPIO]' in obs:
                            parts = obs.replace('[EMPLEADO PROPIO]', '').split('|')
                            if parts:
                                nombre_part = parts[0].strip().upper()
                                if nombre_part == nombre_completo:
                                    if emp_placas:
                                        for p in parts:
                                            if 'Placas:' in p:
                                                placas_part = p.split('Placas:')[1].strip().split('|')[0].strip().upper()
                                                if placas_part == emp_placas:
                                                    entrada_pendiente = reg
                                                    break
                                    else:
                                        entrada_pendiente = reg
                                    break

                if entrada_pendiente:
                    data['entrada_asociada'] = entrada_pendiente.id
                else:
                    return Response(
                        {'empleado_propio': 'Este empleado no tiene un registro de entrada activo con vehículo propio. Primero registre la entrada.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            elif tipo_entidad == 'empleado':
                evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None
                if not evidencia_file:
                    return Response(
                        {'evidencia_fotografica': 'La evidencia fotográfica es obligatoria para la salida de empleados.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

        evidencia_file = request.FILES.get('evidencia_fotografica') if hasattr(request, 'FILES') else None

        serializer = RegistroAccesoSerializer(
            data=data,
            context={'request': request}
        )

        if serializer.is_valid():
            try:
                registro = serializer.save(guardia=request.user)
            except Exception as e:
                print(f"[ERROR] serializer.save() failed: {type(e).__name__}: {str(e)}", file=sys.stderr)
                import traceback
                traceback.print_exc(file=sys.stderr)
                return Response(
                    {'error': f'Error al guardar: {type(e).__name__}: {str(e)}'},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

            if evidencia_file:
                registro.evidencia_fotografica = evidencia_file
                registro.tiene_evidencia = True
                registro.save(update_fields=['evidencia_fotografica', 'tiene_evidencia'])

            if tipo_movimiento == 'entrada':
                if tipo_entidad == 'tracto' and vehiculo_id:
                    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
                    vehiculo.en_instalacion = True
                    vehiculo.conductor_actual_id = conductor_id
                    vehiculo.save(update_fields=['en_instalacion', 'conductor_actual'])

                    # CREAR HISTORIAL DE USO
                    HistorialUsoVehiculo.objects.create(
                        vehiculo=vehiculo,
                        conductor_id=conductor_id,
                        tipo_movimiento='entrada',
                        tipo_entidad=tipo_entidad,
                        turno_id=turno_id,
                        registro_acceso=registro,
                        dentro_instalacion=True
                    )

                elif tipo_entidad == 'empleado' and vehiculo_id:
                    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
                    vehiculo.en_instalacion = True
                    vehiculo.ultimo_empleado_id = empleado_id
                    vehiculo.save(update_fields=['en_instalacion', 'ultimo_empleado'])

                    # CREAR HISTORIAL DE USO
                    HistorialUsoVehiculo.objects.create(
                        vehiculo=vehiculo,
                        empleado_id=empleado_id,
                        tipo_movimiento='entrada',
                        tipo_entidad=tipo_entidad,
                        turno_id=turno_id,
                        registro_acceso=registro,
                        dentro_instalacion=True
                    )

                elif tipo_entidad == 'empleado_propio':
                    data['conductor_pendiente_salida'] = True

            elif tipo_movimiento == 'salida':
                if tipo_entidad == 'tracto' and vehiculo_id and conductor_id:
                    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
                    vehiculo.conductor_actual = None
                    vehiculo.en_instalacion = False
                    vehiculo.save(update_fields=['conductor_actual', 'en_instalacion'])

                    # CREAR HISTORIAL DE SALIDA
                    HistorialUsoVehiculo.objects.create(
                        vehiculo=vehiculo,
                        conductor_id=conductor_id,
                        tipo_movimiento='salida',
                        tipo_entidad=tipo_entidad,
                        turno_id=turno_id,
                        registro_acceso=registro,
                        dentro_instalacion=False
                    )

                    RegistroAcceso.objects.filter(
                        conductor_id=conductor_id,
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).update(conductor_pendiente_salida=False)

                elif tipo_entidad == 'empleado' and vehiculo_id:
                    vehiculo = Vehiculo.objects.get(id=vehiculo_id)
                    vehiculo.en_instalacion = False
                    vehiculo.ultimo_empleado = None
                    vehiculo.save(update_fields=['en_instalacion', 'ultimo_empleado'])

                    # CREAR HISTORIAL DE SALIDA
                    HistorialUsoVehiculo.objects.create(
                        vehiculo=vehiculo,
                        empleado_id=empleado_id,
                        tipo_movimiento='salida',
                        tipo_entidad=tipo_entidad,
                        turno_id=turno_id,
                        registro_acceso=registro,
                        dentro_instalacion=False
                    )

                elif tipo_entidad == 'empleado_propio':
                    RegistroAcceso.objects.filter(
                        tipo_entidad='empleado_propio',
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).update(conductor_pendiente_salida=False)

                elif tipo_entidad == 'conductor' and conductor_id:
                    salida_obs = registro.observaciones or ''
                    entrada_pendiente = RegistroAcceso.objects.filter(
                        conductor_id=conductor_id,
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).first()
                    if entrada_pendiente:
                        obs_actual = entrada_pendiente.observaciones or ''
                        placa_info = f" | Placa salida: {conductor_placa_salida}" if conductor_placa_salida else ''
                        marca_info = f" | Marca salida: {conductor_marca_salida}" if conductor_marca_salida else ''
                        salida_info = f" | SALIDA: {salida_obs}" if salida_obs else ''
                        entrada_pendiente.observaciones = obs_actual + placa_info + marca_info + salida_info
                        entrada_pendiente.save(update_fields=['observaciones'])
                    RegistroAcceso.objects.filter(
                        conductor_id=conductor_id,
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).update(conductor_pendiente_salida=False)
                    conductor_obj = Conductor.objects.filter(id=conductor_id, vehiculo__isnull=False).first()
                    if conductor_obj and conductor_obj.vehiculo:
                        conductor_obj.vehiculo.en_instalacion = False
                        conductor_obj.vehiculo.save(update_fields=['en_instalacion'])

                elif tipo_entidad == 'conductor' and not conductor_id:
                    if registro.entrada_asociada:
                        entrada_pendiente = registro.entrada_asociada
                        obs_actual = entrada_pendiente.observaciones or ''
                        placa_info = f" | Placa salida: {conductor_placa_salida}" if conductor_placa_salida else ''
                        marca_info = f" | Marca salida: {conductor_marca_salida}" if conductor_marca_salida else ''
                        entrada_pendiente.observaciones = obs_actual + placa_info + marca_info
                        entrada_pendiente.conductor_pendiente_salida = False
                        entrada_pendiente.save(update_fields=['observaciones', 'conductor_pendiente_salida'])
                    RegistroAcceso.objects.filter(
                        tipo_entidad='conductor',
                        tipo_movimiento='entrada',
                        conductor_pendiente_salida=True
                    ).exclude(id=registro.entrada_asociada_id if registro.entrada_asociada else None).update(
                        conductor_pendiente_salida=False
                    )

                elif tipo_entidad == 'visitante' and registro.entrada_asociada:
                    entrada_pendiente = registro.entrada_asociada
                    obs_actual = entrada_pendiente.observaciones or ''
                    placa_info = f" | Placa salida: {visitante_placa_salida}" if visitante_placa_salida else ''
                    entrada_pendiente.observaciones = obs_actual + placa_info
                    entrada_pendiente.conductor_pendiente_salida = False
                    entrada_pendiente.save(update_fields=['observaciones', 'conductor_pendiente_salida'])

            registrar_auditoria(
                usuario=request.user,
                modulo='accesos',
                accion='crear',
                descripcion=f'Se creó registro de acceso {registro.id} ({registro.tipo_movimiento} - {registro.tipo_entidad}).',
                entidad_tipo='RegistroAcceso',
                entidad_id=registro.id,
                turno=registro.turno
            )

            return Response(
                {
                    'message': 'Registro de acceso creado correctamente.',
                    'data': RegistroAccesoSerializer(
                        registro,
                        context={'request': request}
                    ).data
                },
                status=status.HTTP_201_CREATED
            )

        print(f"[DEBUG] serializer.errors: {serializer.errors}", file=sys.stderr)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class VehiculosEnInstalacionAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tipo = request.query_params.get('tipo')

        vehiculos = Vehiculo.objects.filter(en_instalacion=True)

        if tipo == 'tracto':
            vehiculos = vehiculos.filter(tipo_entidad='tracto')
        elif tipo == 'empleado':
            vehiculos = vehiculos.filter(tipo_entidad='empleado')

        vehiculos = vehiculos.select_related('conductor_actual', 'ultimo_empleado').order_by('placa')

        data = []
        for v in vehiculos:
            conductor_info = None
            if v.conductor_actual:
                conductor_info = {
                    'id': v.conductor_actual.id,
                    'nombre_completo': v.conductor_actual.nombre_completo
                }
            empleado_info = None
            if v.ultimo_empleado:
                empleado_info = {
                    'id': v.ultimo_empleado.id,
                    'nombre_completo': v.ultimo_empleado.nombre_completo,
                    'numero_empleado': v.ultimo_empleado.numero_empleado
                }

            data.append({
                'id': v.id,
                'placa': v.placa,
                'clave_interna': v.clave_interna,
                'tipo_entidad': v.tipo_entidad,
                'marca': v.marca,
                'modelo': v.modelo,
                'conductor_actual': v.conductor_actual_id,
                'conductor_actual_info': conductor_info,
                'ultimo_empleado': v.ultimo_empleado_id,
                'ultimo_empleado_info': empleado_info,
            })

        return Response(data)


class VehiculosDisponiblesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tipo_movimiento = request.query_params.get('tipo_movimiento')
        tipo_vehiculo = request.query_params.get('tipo')

        vehiculos = Vehiculo.objects.all()

        if tipo_vehiculo == 'tracto':
            vehiculos = vehiculos.filter(tipo_entidad='tracto')
            if tipo_movimiento == 'entrada':
                vehiculos = vehiculos.filter(
                    models.Q(en_instalacion=False) |
                    models.Q(en_instalacion=True, conductor_actual__isnull=True)
                )
            elif tipo_movimiento == 'salida':
                vehiculos = vehiculos.filter(
                    en_instalacion=True,
                    conductor_actual__isnull=False
                )
        elif tipo_vehiculo == 'empleado':
            vehiculos = vehiculos.filter(tipo_entidad='empleado')
            if tipo_movimiento == 'entrada':
                vehiculos = vehiculos.filter(en_instalacion=False)
            elif tipo_movimiento == 'salida':
                vehiculos = vehiculos.filter(en_instalacion=True)
        else:
            if tipo_movimiento == 'entrada':
                vehiculos = vehiculos.filter(en_instalacion=False)
            elif tipo_movimiento == 'salida':
                vehiculos = vehiculos.filter(en_instalacion=True)

        vehiculos = vehiculos.select_related('conductor_actual', 'ultimo_empleado').order_by('placa')

        data = []
        for v in vehiculos:
            conductor_info = None
            if v.conductor_actual:
                conductor_info = {
                    'id': v.conductor_actual.id,
                    'nombre_completo': v.conductor_actual.nombre_completo
                }

            data.append({
                'id': v.id,
                'placa': v.placa,
                'clave_interna': v.clave_interna,
                'tipo_entidad': v.tipo_entidad,
                'marca': v.marca,
                'modelo': v.modelo,
                'conductor_actual': v.conductor_actual_id,
                'conductor_actual_info': conductor_info,
                'en_instalacion': v.en_instalacion,
            })

        return Response(data)


class EmpleadosConVehiculoDisponibleAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tipo_movimiento = request.query_params.get('tipo_movimiento')

        asignaciones = AsignacionVehiculoEmpleado.objects.filter(
            activa=True
        ).select_related('vehiculo', 'empleado').order_by('empleado__nombre_completo')

        data = []
        for a in asignaciones:
            if not a.vehiculo:
                continue

            vehiculo_esta_dentro = a.vehiculo.en_instalacion

            if tipo_movimiento == 'entrada' and vehiculo_esta_dentro:
                continue
            if tipo_movimiento == 'salida' and not vehiculo_esta_dentro:
                continue

            data.append({
                'id': a.empleado.id,
                'numero_empleado': a.empleado.numero_empleado,
                'nombre_completo': a.empleado.nombre_completo,
                'departamento': a.empleado.departamento,
                'puesto': a.empleado.puesto,
                'vehiculo_id': a.vehiculo.id,
                'vehiculo_placa': a.vehiculo.placa,
                'vehiculo_clave_interna': a.vehiculo.clave_interna,
                'vehiculo_marca': a.vehiculo.marca,
                'vehiculo_modelo': a.vehiculo.modelo,
                'vehiculo_en_instalacion': vehiculo_esta_dentro,
                'asignacion_id': a.id,
            })

        return Response(data)


class ConductoresDisponiblesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        tipo_movimiento = request.query_params.get('tipo_movimiento')
        vehiculo_id = request.query_params.get('vehiculo_id')

        if tipo_movimiento == 'entrada':
            # Usar nueva tabla de asignaciones
            from platform_core.models import AsignacionConductorVehiculo

            asignaciones_activas = AsignacionConductorVehiculo.objects.filter(
                activa=True,
                fecha_desasignacion__isnull=True
            ).select_related('conductor', 'vehiculo')

            if vehiculo_id:
                try:
                    asignaciones_activas = asignaciones_activas.filter(vehiculo_id=int(vehiculo_id))
                except (ValueError, TypeError):
                    pass

            data = []
            for asignacion in asignaciones_activas:
                c = asignacion.conductor
                tiene_entrada_pendiente = RegistroAcceso.objects.filter(
                    conductor=c,
                    tipo_movimiento='entrada',
                    conductor_pendiente_salida=True
                ).exists()
                if tiene_entrada_pendiente:
                    continue
                data.append({
                    'id': c.id,
                    'nombre_completo': c.nombre_completo,
                    'licencia': c.licencia,
                    'vehiculo_id': asignacion.vehiculo_id,
                    'vehiculo_placa': asignacion.vehiculo.placa,
                    'vehiculo_marca': asignacion.vehiculo.marca,
                    'vehiculo_modelo': asignacion.vehiculo.modelo,
                    'asignacion_id': asignacion.id,
                })
        elif tipo_movimiento == 'salida':
            from platform_core.models import AsignacionConductorVehiculo

            asignaciones_activas = AsignacionConductorVehiculo.objects.filter(
                activa=True,
                fecha_desasignacion__isnull=True
            ).select_related('conductor', 'vehiculo')

            if vehiculo_id:
                try:
                    asignaciones_activas = asignaciones_activas.filter(vehiculo_id=int(vehiculo_id))
                except (ValueError, TypeError):
                    pass

            data = []
            for asignacion in asignaciones_activas:
                c = asignacion.conductor
                tiene_entrada_pendiente = RegistroAcceso.objects.filter(
                    conductor=c,
                    tipo_movimiento='entrada',
                    conductor_pendiente_salida=True
                ).exists()
                if not tiene_entrada_pendiente:
                    continue
                data.append({
                    'id': c.id,
                    'nombre_completo': c.nombre_completo,
                    'licencia': c.licencia,
                    'vehiculo_id': asignacion.vehiculo_id,
                    'vehiculo_placa': asignacion.vehiculo.placa,
                    'vehiculo_marca': asignacion.vehiculo.marca,
                    'vehiculo_modelo': asignacion.vehiculo.modelo,
                    'asignacion_id': asignacion.id,
                })
        else:
            data = []

        return Response(data)


class RegistroAccesoPendientesSalidaAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from platform_core.models import RegistroAcceso, AsignacionVehiculoEmpleado
        from django.db.models import Q

        data = {
            'tractos_pendientes': [],
            'conductores_pendientes': [],
            'empleados_empresa_pendientes': [],
            'empleados_propio_pendientes': [],
            'visitantes_pendientes': []
        }

        tractos_pendientes = RegistroAcceso.objects.filter(
            tipo_entidad='tracto',
            tipo_movimiento='entrada'
        ).filter(
            Q(conductor_pendiente_salida=True) | Q(vehiculo__en_instalacion=True)
        ).exclude(
            id__in=RegistroAcceso.objects.filter(
                tipo_movimiento='salida',
                entrada_asociada__isnull=False
            ).values_list('entrada_asociada_id', flat=True)
        ).select_related('vehiculo', 'conductor', 'turno').distinct()

        seen_vehiculo_ids = set()
        for r in tractos_pendientes:
            if r.vehiculo_id and r.vehiculo_id in seen_vehiculo_ids:
                continue
            if r.vehiculo_id:
                seen_vehiculo_ids.add(r.vehiculo_id)
            data['tractos_pendientes'].append({
                'id': r.id,
                'vehiculo_id': r.vehiculo_id,
                'vehiculo_placa': r.vehiculo.placa if r.vehiculo else None,
                'vehiculo_clave_interna': r.vehiculo.clave_interna if r.vehiculo else None,
                'conductor_id': r.conductor_id,
                'conductor_nombre': r.conductor.nombre_completo if r.conductor else None,
                'fecha_entrada': r.fecha_hora,
                'turno_id': r.turno_id,
                'turno_tipo': r.turno.tipo_turno if r.turno else None,
            })

        conductores_pendientes = RegistroAcceso.objects.filter(
            conductor_pendiente_salida=True,
            conductor_id__isnull=False,
            tipo_entidad__in=('conductor', 'tracto')
        ).exclude(
            id__in=RegistroAcceso.objects.filter(
                tipo_movimiento='salida',
                entrada_asociada__isnull=False
            ).values_list('entrada_asociada_id', flat=True)
        ).exclude(
            tipo_entidad='visitante'
        ).select_related('conductor', 'turno')

        seen_conductor_ids = set()
        for r in conductores_pendientes:
            if r.conductor_id in seen_conductor_ids:
                continue
            seen_conductor_ids.add(r.conductor_id)

            obs = r.observaciones or ''
            conductor_nombre = r.conductor.nombre_completo if r.conductor else None
            conductor_placa = None
            conductor_marca = None

            if '[CONDUCTOR]' in obs:
                parts = obs.replace('[CONDUCTOR]', '').split('|')
                if parts and not conductor_nombre:
                    conductor_nombre = parts[0].strip()
                for p in parts:
                    if 'Placas:' in p:
                        conductor_placa = p.split('Placas:')[1].strip().split('|')[0].strip()
                    if 'Marca:' in p:
                        conductor_marca = p.split('Marca:')[1].strip()

            data['conductores_pendientes'].append({
                'id': r.id,
                'conductor_id': r.conductor_id,
                'conductor_nombre': conductor_nombre,
                'conductor_placa': conductor_placa,
                'conductor_marca': conductor_marca,
                'observaciones': obs,
                'fecha_entrada': r.fecha_hora,
                'turno_id': r.turno_id,
                'turno_tipo': r.turno.tipo_turno if r.turno else None,
                'vehiculo_id': r.vehiculo_id,
                'tipo_entidad': r.tipo_entidad,
                'tipo_conductor': 'conductor_sin_vehiculo',
            })

        empleados_empresa_pendientes = RegistroAcceso.objects.filter(
            tipo_entidad='empleado',
            tipo_movimiento='entrada',
            vehiculo__isnull=False,
            vehiculo__en_instalacion=True
        ).exclude(
            id__in=RegistroAcceso.objects.filter(
                tipo_movimiento='salida',
                entrada_asociada__isnull=False
            ).values_list('entrada_asociada_id', flat=True)
        ).select_related('empleado', 'vehiculo', 'turno').distinct()

        for r in empleados_empresa_pendientes:
            data['empleados_empresa_pendientes'].append({
                'id': r.id,
                'vehiculo_id': r.vehiculo_id,
                'vehiculo_placa': r.vehiculo.placa if r.vehiculo else None,
                'vehiculo_clave_interna': r.vehiculo.clave_interna if r.vehiculo else None,
                'vehiculo_marca': r.vehiculo.marca if r.vehiculo else None,
                'vehiculo_modelo': r.vehiculo.modelo if r.vehiculo else None,
                'vehiculo_color': r.vehiculo.color if r.vehiculo else None,
                'empleado_id': r.empleado_id,
                'empleado_nombre': r.empleado.nombre_completo if r.empleado else None,
                'empleado_numero': r.empleado.numero_empleado if r.empleado else None,
                'fecha_entrada': r.fecha_hora,
                'turno_id': r.turno_id,
                'turno_tipo': r.turno.tipo_turno if r.turno else None,
            })

        empleados_propio_pendientes = RegistroAcceso.objects.filter(
            tipo_entidad='empleado',
            tipo_movimiento='entrada',
            conductor_pendiente_salida=True
        ).exclude(
            id__in=RegistroAcceso.objects.filter(
                tipo_movimiento='salida',
                entrada_asociada__isnull=False
            ).values_list('entrada_asociada_id', flat=True)
        ).exclude(
            vehiculo__isnull=False
        ).select_related('turno').distinct()

        for r in empleados_propio_pendientes:
            obs = r.observaciones or ''
            emp_nombre = None
            emp_placas = None
            emp_marca = None

            if '[EMPLEADO PROPIO]' in obs:
                parts = obs.replace('[EMPLEADO PROPIO]', '').split('|')
                if parts:
                    emp_nombre = parts[0].strip()
                for p in parts:
                    if 'Placas:' in p:
                        emp_placas = p.split('Placas:')[1].strip().split('|')[0].strip()
                    if 'Marca:' in p:
                        emp_marca = p.split('Marca:')[1].strip()

            data['empleados_propio_pendientes'].append({
                'id': r.id,
                'empleado_nombre': emp_nombre,
                'empleado_placas': emp_placas,
                'empleado_marca': emp_marca,
                'observaciones': obs,
                'fecha_entrada': r.fecha_hora,
                'turno_id': r.turno_id,
                'turno_tipo': r.turno.tipo_turno if r.turno else None,
})

        visitantes_pendientes = RegistroAcceso.objects.filter(
            tipo_entidad='visitante',
            tipo_movimiento='entrada',
            conductor_pendiente_salida=True
        ).exclude(
            id__in=RegistroAcceso.objects.filter(
                tipo_movimiento='salida',
                entrada_asociada__isnull=False
            ).values_list('entrada_asociada_id', flat=True)
).select_related('visitante', 'turno', 'vehiculo').distinct()

        for r in visitantes_pendientes:
            obs = r.observaciones or ''
            visitante_nombre = None
            visitante_placas = None

            if '[VISITANTE]' in obs:
                parts = obs.replace('[VISITANTE]', '').split('|')
                if parts:
                    visitante_nombre = parts[0].strip()
                for p in parts:
                    if 'Placas:' in p:
                        visitante_placas = p.split('Placas:')[1].strip().split('|')[0].split('Motivo:')[0].strip()
                        break

            data['visitantes_pendientes'].append({
                'id': r.id,
                'visitante_id': r.visitante_id,
                'visitante_nombre': visitante_nombre,
                'visitante_placas': visitante_placas,
                'vehiculo_id': r.vehiculo_id,
                'vehiculo_placa': r.vehiculo.placa if r.vehiculo else None,
                'fecha_entrada': r.fecha_hora,
                'turno_id': r.turno_id,
                'turno_tipo': r.turno.tipo_turno if r.turno else None,
            })

        return Response(data)


class VisitantesPendientesSalidaAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from platform_core.models import RegistroAcceso

        visitantes_pendientes = RegistroAcceso.objects.filter(
            tipo_entidad='visitante',
            tipo_movimiento='entrada'
        ).exclude(
            id__in=RegistroAcceso.objects.filter(
                tipo_movimiento='salida',
                entrada_asociada__isnull=False
            ).values_list('entrada_asociada_id', flat=True)
        ).select_related('visitante', 'turno', 'vehiculo').order_by('-fecha_hora')

        data = []
        for r in visitantes_pendientes:
            obs = r.observaciones or ''
            visitante_nombre = None
            visitante_placas = None

            if '[VISITANTE]' in obs:
                parts = obs.replace('[VISITANTE]', '').split('|')
                if parts:
                    visitante_nombre = parts[0].strip()
                for p in parts:
                    if 'Placas:' in p:
                        visitante_placas = p.split('Placas:')[1].strip().split('|')[0].split('Motivo:')[0].strip()
                        break

            data.append({
                'id': r.id,
                'entrada_asociada_id': r.id,
                'visitante_id': r.visitante_id,
                'visitante_nombre': visitante_nombre,
                'visitante_placas': visitante_placas,
                'vehiculo_placa': r.vehiculo.placa if r.vehiculo else None,
                'vehiculo_clave_interna': r.vehiculo.clave_interna if r.vehiculo else None,
                'fecha_entrada': r.fecha_hora,
                'turno_id': r.turno_id,
                'turno_tipo': r.turno.tipo_turno if r.turno else None,
            })

        return Response(data)


# =========================
# API VIEWS - CHECKLIST TRACTO
# =========================

class ChecklistTractoItemCatalogoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ChecklistTractoItemCatalogoSerializer

    def get_queryset(self):
        queryset = ChecklistTractoItemCatalogo.objects.filter(
            activo=True
        ).order_by('seccion', 'orden', 'id')

        seccion = self.request.query_params.get('seccion')

        if seccion:
            queryset = queryset.filter(seccion=seccion)

        return queryset


class ChecklistTractoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ChecklistTractoSerializer

    def get_queryset(self):
        queryset = ChecklistTracto.objects.select_related(
            'registro_acceso',
            'turno',
            'guardia',
            'vehiculo',
            'conductor'
        ).prefetch_related(
            'resultados__item',
            'llantas',
            'evidencias'
        ).all().order_by('-fecha_hora', '-id')

        turno = self.request.query_params.get('turno')
        guardia = self.request.query_params.get('guardia')
        vehiculo = self.request.query_params.get('vehiculo')
        estatus_general = self.request.query_params.get('estatus_general')

        if turno:
            queryset = queryset.filter(turno_id=turno)

        if guardia:
            queryset = queryset.filter(guardia_id=guardia)

        if vehiculo:
            queryset = queryset.filter(vehiculo_id=vehiculo)

        if estatus_general:
            queryset = queryset.filter(estatus_general=estatus_general)

        return queryset


class ChecklistTractoCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        import sys
        import traceback

        print(f"[DEBUG] === ChecklistTractoCreateAPIView called ===", file=sys.stderr)
        print(f"[DEBUG] request.data keys: {list(request.data.keys())}", file=sys.stderr)
        for key in request.data.keys():
            val = request.data.get(key)
            val_type = type(val).__name__
            if hasattr(val, 'size'):
                print(f"[DEBUG]   {key}: {val_type} (size={val.size}, name={val.name})", file=sys.stderr)
            elif isinstance(val, str) and len(val) > 100:
                print(f"[DEBUG]   {key}: {val_type} (len={len(val)}, preview={val[:50]}...)", file=sys.stderr)
            else:
                print(f"[DEBUG]   {key}: {val_type} = {val}", file=sys.stderr)

        try:
            raw_data = request.data
            if hasattr(raw_data, 'lists'):
                data = {k: raw_data.getlist(k)[0] if len(raw_data.getlist(k)) == 1 else raw_data.getlist(k) for k in raw_data.keys()}
            elif hasattr(raw_data, 'dict'):
                data = raw_data.dict()
            else:
                data = dict(raw_data)

            from platform_core.models import RegistroAcceso

            tipo_movimiento = data.get('tipo_movimiento')
            vehiculo_id = data.get('vehiculo')
            turno_id = data.get('turno')

            if tipo_movimiento == 'salida' and vehiculo_id and turno_id:
                entrada_existe = RegistroAcceso.objects.filter(
                    turno_id=turno_id,
                    vehiculo_id=vehiculo_id,
                    tipo_movimiento='entrada'
                ).exists()
                if not entrada_existe:
                    return Response(
                        {'vehiculo': 'Este vehículo no tiene un registro de entrada en este turno. Primero registre la entrada.'},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if isinstance(data.get('resultados'), str):
                try:
                    data['resultados'] = json.loads(data['resultados'])
                except Exception:
                    return Response(
                        {'resultados': ['El formato de resultados no es válido.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            if isinstance(data.get('llantas'), str):
                try:
                    data['llantas'] = json.loads(data['llantas'])
                except Exception:
                    return Response(
                        {'llantas': ['El formato de llaves no es válido.']},
                        status=status.HTTP_400_BAD_REQUEST
                    )

            serializer = ChecklistTractoCreateSerializer(
                data=data,
                context={'request': request}
            )

            print(f"[DEBUG] serializer.is_valid() = {serializer.is_valid()}", file=sys.stderr)
            if not serializer.is_valid():
                print(f"[DEBUG] serializer.errors: {serializer.errors}", file=sys.stderr)
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            checklist = serializer.save()
            print(f"[DEBUG] checklist saved: id={checklist.id}", file=sys.stderr)

            registrar_auditoria(
                usuario=request.user,
                modulo='checklist_tracto',
                accion='crear',
                descripcion=f'Se creó checklist de tracto {checklist.id} para la unidad {checklist.vehiculo.placa if checklist.vehiculo else ""}.',
                entidad_tipo='ChecklistTracto',
                entidad_id=checklist.id,
                turno=checklist.turno
            )

            return Response(
                {
                    'message': 'Checklist de tracto creado correctamente.',
                    'data': ChecklistTractoSerializer(
                        checklist,
                        context={'request': request}
                    ).data
                },
                status=status.HTTP_201_CREATED
            )

        except Exception as e:
            print(f"[ERROR] UNEXPECTED ERROR: {e}", file=sys.stderr)
            traceback.print_exc()
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ChecklistTractoDetailAPIView(generics.RetrieveAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = ChecklistTractoSerializer
    queryset = ChecklistTracto.objects.select_related(
        'registro_acceso',
        'turno',
        'guardia',
        'vehiculo',
        'conductor'
    ).prefetch_related(
        'resultados__item',
        'llantas',
        'evidencias'
    ).all()


# =========================
# API VIEWS - AUDITORÍA / REPORTES
# =========================

class AuditLogListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = AuditLogSerializer

    def get_queryset(self):
        queryset = AuditLog.objects.select_related(
            'usuario',
            'turno'
        ).all().order_by('-fecha_hora', '-id')

        modulo = self.request.query_params.get('modulo')
        accion = self.request.query_params.get('accion')
        turno = self.request.query_params.get('turno')
        usuario = self.request.query_params.get('usuario')

        if modulo:
            queryset = queryset.filter(modulo=modulo)

        if accion:
            queryset = queryset.filter(accion=accion)

        if turno:
            queryset = queryset.filter(turno_id=turno)

        if usuario:
            queryset = queryset.filter(usuario_id=usuario)

        return queryset


class NotificacionListCreateAPIView(generics.ListCreateAPIView):
    permission_classes = [IsAuthenticated]
    serializer_class = NotificacionSerializer

    def get_queryset(self):
        return Notificacion.objects.filter(
            usuario=self.request.user
        ).order_by('-fecha_hora', '-id')

    def create(self, request, *args, **kwargs):
        serializer = NotificacionCreateSerializer(data=request.data)
        if serializer.is_valid():
            notificacion = Notificacion.objects.create(
                usuario=request.user,
                titulo=serializer.validated_data['titulo'],
                mensaje=serializer.validated_data['mensaje'],
                tipo=serializer.validated_data.get('tipo', 'info'),
                link=serializer.validated_data.get('link', '')
            )
            return Response(
                NotificacionSerializer(notificacion).data,
                status=status.HTTP_201_CREATED
            )
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class NotificacionMarkReadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request, pk):
        try:
            notificacion = Notificacion.objects.get(pk=pk, usuario=request.user)
        except Notificacion.DoesNotExist:
            return Response(
                {'error': 'Notificación no encontrada'},
                status=status.HTTP_404_NOT_FOUND
            )

        notificacion.leida = True
        notificacion.save()

        return Response({'message': 'Notificación marcada como leída'})


class NotificacionMarkAllReadAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def patch(self, request):
        Notificacion.objects.filter(
            usuario=request.user,
            leida=False
        ).update(leida=True)

        return Response({'message': 'Todas las notificaciones marcadas como leídas'})


class EnviarNotificacionTurnoAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def post(self, request):
        tipo = request.data.get('tipo')
        turno_id = request.data.get('turno_id')

        if tipo not in ['turno_abierto', 'turno_cerrado']:
            return Response(
                {'error': 'Tipo de notificación inválido'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            turno = Turno.objects.get(pk=turno_id)
        except Turno.DoesNotExist:
            return Response(
                {'error': 'Turno no encontrado'},
                status=status.HTTP_404_NOT_FOUND
            )

        if tipo == 'turno_abierto':
            titulo = 'Turno abierto'
            mensaje = f'Se ha abierto el turno {turno.get_tipo_turno_display()} ({turno.fecha})'
        else:
            titulo = 'Turno cerrado'
            mensaje = f'Se ha cerrado el turno {turno.get_tipo_turno_display()} ({turno.fecha})'

        all_users = User.objects.filter(is_active=True, profile__is_active_user=True)

        for user in all_users:
            Notificacion.objects.create(
                usuario=user,
                titulo=titulo,
                mensaje=mensaje,
                tipo=tipo,
                link=f'/turnos/{turno_id}'
            )

        return Response({
            'message': f'Notificación enviada a {all_users.count()} usuarios',
            'total': all_users.count()
        })


class ReportesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        total_registros = RegistroAcceso.objects.count()
        total_checklists_tracto = ChecklistTracto.objects.count()
        total_turnos_abiertos = Turno.objects.filter(abierto=True).count()
        total_vehiculos = Vehiculo.objects.count()

        registros_por_tipo = list(
            RegistroAcceso.objects.values('tipo_entidad')
            .annotate(total=Count('id'))
            .order_by('tipo_entidad')
        )

        movimientos_por_tipo = list(
            RegistroAcceso.objects.values('tipo_movimiento')
            .annotate(total=Count('id'))
            .order_by('tipo_movimiento')
        )

        checklists_por_estatus = list(
            ChecklistTracto.objects.values('estatus_general')
            .annotate(total=Count('id'))
            .order_by('estatus_general')
        )

        return Response({
            'total_registros': total_registros,
            'total_checklists_tracto': total_checklists_tracto,
            'total_turnos_abiertos': total_turnos_abiertos,
            'total_vehiculos': total_vehiculos,
            'registros_por_tipo': registros_por_tipo,
            'movimientos_por_tipo': movimientos_por_tipo,
            'checklists_por_estatus': checklists_por_estatus,
        }, status=status.HTTP_200_OK)


class ReportePDFAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from io import BytesIO
        from django.conf import settings
        from datetime import datetime

        fecha_inicio = request.GET.get('fecha_inicio')
        fecha_fin = request.GET.get('fecha_fin')

        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape, topMargin=0.5*inch, bottomMargin=0.5*inch)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#059669'),
            spaceAfter=20,
        )

        elements = []
        elements.append(Paragraph('LRA - Reporte de Actividad', title_style))
        elements.append(Spacer(1, 0.3*inch))

        fecha_texto = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
        if fecha_inicio and fecha_fin:
            fecha_texto += f' | Periodo: {fecha_inicio} al {fecha_fin}'
        elements.append(Paragraph(fecha_texto, styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))

        total_registros = RegistroAcceso.objects.count()
        total_checklists = ChecklistTracto.objects.count()
        total_turnos = Turno.objects.count()
        total_vehiculos = Vehiculo.objects.count()
        turnos_abiertos = Turno.objects.filter(abierto=True).count()

        summary_data = [
            ['Metric', 'Total'],
            ['Total Registros', str(total_registros)],
            ['Total Checklists Tracto', str(total_checklists)],
            ['Total Turnos', str(total_turnos)],
            ['Turnos Abiertos', str(turnos_abiertos)],
            ['Total Vehículos', str(total_vehiculos)],
        ]

        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#059669')),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))

        tipo_data = [['Tipo Entidad', 'Total']]
        for item in RegistroAcceso.objects.values('tipo_entidad').annotate(total=Count('id')).order_by('tipo_entidad'):
            tipo_data.append([item['tipo_entidad'], str(item['total'])])
        tipo_table = Table(tipo_data, colWidths=[3*inch, 2*inch])
        tipo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ]))
        elements.append(Paragraph('Registros por Tipo', styles['Heading2']))
        elements.append(tipo_table)
        elements.append(Spacer(1, 0.3*inch))

        checklist_data = [['Estatus', 'Total']]
        for item in ChecklistTracto.objects.values('estatus_general').annotate(total=Count('id')).order_by('estatus_general'):
            checklist_data.append([item['estatus_general'] or 'Sin estatus', str(item['total'])])
        checklist_table = Table(checklist_data, colWidths=[3*inch, 2*inch])
        checklist_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#059669')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#e5e7eb')),
        ]))
        elements.append(Paragraph('Checklists por Estatus', styles['Heading2']))
        elements.append(checklist_table)

        doc.build(elements)

        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="reporte_lra_{datetime.now().strftime("%Y%m%d_%H%M")}.pdf"'
        return response


class ReporteExcelAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime

        wb = Workbook()

        header_fill = PatternFill(start_color='059669', end_color='059669', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        ws_summary = wb.active
        ws_summary.title = 'Resumen'

        ws_summary['A1'] = 'LRA - Reporte de Actividad'
        ws_summary['A1'].font = Font(size=16, bold=True, color='059669')
        ws_summary['A2'] = f'Generado: {datetime.now().strftime("%d/%m/%Y %H:%M")}'

        ws_summary['A4'] = 'Métrica'
        ws_summary['B4'] = 'Total'
        ws_summary['A4'].fill = header_fill
        ws_summary['B4'].fill = header_fill
        ws_summary['A4'].font = header_font
        ws_summary['B4'].font = header_font

        summary_data = [
            ('Total Registros', RegistroAcceso.objects.count()),
            ('Total Checklists Tracto', ChecklistTracto.objects.count()),
            ('Total Turnos', Turno.objects.count()),
            ('Turnos Abiertos', Turno.objects.filter(abierto=True).count()),
            ('Total Vehículos', Vehiculo.objects.count()),
        ]

        for i, (metric, value) in enumerate(summary_data, start=5):
            ws_summary[f'A{i}'] = metric
            ws_summary[f'B{i}'] = value

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15

        ws_registros = wb.create_sheet('Registros')
        ws_registros['A1'] = 'Registros por Tipo'
        ws_registros['A1'].font = Font(size=14, bold=True, color='059669')
        ws_registros['A2'] = 'Tipo Entidad'
        ws_registros['B2'] = 'Total'
        ws_registros['A2'].fill = header_fill
        ws_registros['B2'].fill = header_fill
        ws_registros['A2'].font = header_font
        ws_registros['B2'].font = header_font

        row = 3
        for item in RegistroAcceso.objects.values('tipo_entidad').annotate(total=Count('id')).order_by('tipo_entidad'):
            ws_registros[f'A{row}'] = item['tipo_entidad']
            ws_registros[f'B{row}'] = item['total']
            row += 1

        ws_registros.column_dimensions['A'].width = 20
        ws_registros.column_dimensions['B'].width = 15

        ws_checklists = wb.create_sheet('Checklists')
        ws_checklists['A1'] = 'Checklists por Estatus'
        ws_checklists['A1'].font = Font(size=14, bold=True, color='059669')
        ws_checklists['A2'] = 'Estatus'
        ws_checklists['B2'] = 'Total'
        ws_checklists['A2'].fill = header_fill
        ws_checklists['B2'].fill = header_fill
        ws_checklists['A2'].font = header_font
        ws_checklists['B2'].font = header_font

        row = 3
        for item in ChecklistTracto.objects.values('estatus_general').annotate(total=Count('id')).order_by('estatus_general'):
            ws_checklists[f'A{row}'] = item['estatus_general'] or 'Sin estatus'
            ws_checklists[f'B{row}'] = item['total']
            row += 1

        ws_checklists.column_dimensions['A'].width = 20
        ws_checklists.column_dimensions['B'].width = 15

        ws_vehiculos = wb.create_sheet('Vehículos')
        ws_vehiculos['A1'] = 'Vehículos'
        ws_vehiculos['A1'].font = Font(size=14, bold=True, color='059669')
        headers_v = ['Placa', 'Tipo Entidad', 'Categoría', 'Marca', 'Modelo', 'Activo']
        for col, header in enumerate(headers_v, start=1):
            cell = ws_vehiculos.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        row = 3
        for vehiculo in Vehiculo.objects.all()[:100]:
            ws_vehiculos[f'A{row}'] = vehiculo.placa
            ws_vehiculos[f'B{row}'] = vehiculo.tipo_entidad
            ws_vehiculos[f'C{row}'] = vehiculo.categoria
            ws_vehiculos[f'D{row}'] = vehiculo.marca
            ws_vehiculos[f'E{row}'] = vehiculo.modelo
            ws_vehiculos[f'F{row}'] = 'Sí' if vehiculo.activo else 'No'
            row += 1

        for col in range(1, 7):
            ws_vehiculos.column_dimensions[get_column_letter(col)].width = 15

        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_lra_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx"'

        wb.save(response)
        return response


# =========================
# WEB - LOGIN
# =========================

def web_login_view(request):
    from django.contrib.auth import authenticate, login

    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')

        user = authenticate(username=username, password=password)

        if user is not None:
            login(request, user)
            # Verificar si es admin
            try:
                if user.is_superuser:
                    return redirect('/web/dashboard/')
                perfil = getattr(user, 'profile', None)
                if perfil and perfil.role == 'admin':
                    return redirect('/web/dashboard/')
            except Exception:
                pass
            return redirect('/web/guardias/')
        else:
            return render(request, 'accounts/login.html', {
                'error': 'Usuario o contraseña incorrectos.'
            })

    return render(request, 'accounts/login.html')


def web_logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    return redirect('/web/login/')


# =========================
# WEB - DASHBOARD
# =========================

@login_required
def admin_dashboard_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos para acceder al panel de administración.')
        return redirect('web-guardias')

    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta
    from django.contrib.auth import get_user_model
    from django.conf import settings

    User = get_user_model()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    guardia_id = request.GET.get('guardia')

    queryset = RegistroAcceso.objects.all()

    if fecha_inicio:
        queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)
    if guardia_id:
        queryset = queryset.filter(guardia_id=guardia_id)

    total_registros = queryset.count()
    total_turnos_abiertos = Turno.objects.filter(abierto=True).count()
    total_vehiculos = Vehiculo.objects.filter(activo=True).count()
    total_checklists_tracto = ChecklistTracto.objects.count()

    ultimos_registros = queryset.select_related(
        'turno', 'guardia', 'vehiculo', 'empleado', 'conductor', 'visitante'
    ).order_by('-fecha_hora')[:10]

    ultimos_checklists = ChecklistTracto.objects.select_related(
        'vehiculo', 'conductor', 'guardia', 'turno'
    ).order_by('-fecha_hora')[:10]

    seven_days_ago = timezone.now() - timedelta(days=7)

    daily_data = queryset.filter(
        fecha_hora__gte=seven_days_ago
    ).values('fecha_hora__date', 'tipo_movimiento').annotate(
        total=Count('id')
    ).order_by('fecha_hora__date')

    dias = sorted(set(daily_data.values_list('fecha_hora__date', flat=True)))
    labels = [d.strftime('%d/%m') for d in dias]
    entradas_data = []
    salidas_data = []
    for d in dias:
        ent = daily_data.filter(fecha_hora__date=d, tipo_movimiento='entrada').first()
        sal = daily_data.filter(fecha_hora__date=d, tipo_movimiento='salida').first()
        entradas_data.append(ent['total'] if ent else 0)
        salidas_data.append(sal['total'] if sal else 0)

    tipo_data = queryset.values('tipo_entidad').annotate(total=Count('id'))
    tipos_labels = [t['tipo_entidad'] for t in tipo_data]
    tipos_data = [t['total'] for t in tipo_data]

    checklist_data = ChecklistTracto.objects.values('estatus_general').annotate(total=Count('id'))
    estatus_labels = [c['estatus_general'] for c in checklist_data]
    estatus_data = [c['total'] for c in checklist_data]

    guardias = User.objects.all().order_by('username')

    turnos_stats = {
        'total': Turno.objects.count(),
        'abiertos': Turno.objects.filter(abierto=True).count(),
        'cerrados_hoy': Turno.objects.filter(
            abierto=False,
            hora_cierre__date=timezone.now().date()
        ).count(),
    }

    usuarios_stats = {
        'total': User.objects.count(),
        'total_guardias': User.objects.filter(profile__role='guardia').count(),
        'total_admins': User.objects.filter(profile__role='admin').count(),
        'activos': User.objects.filter(is_active=True, profile__is_active_user=True).count(),
        'inactivos': User.objects.filter(is_active=False).count(),
    }

    context = {
        'user': request.user,
        'total_registros': total_registros,
        'total_turnos_abiertos': total_turnos_abiertos,
        'total_vehiculos': total_vehiculos,
        'total_checklists_tracto': total_checklists_tracto,
        'ultimos_registros': ultimos_registros,
        'ultimos_checklists': ultimos_checklists,
        'guardias': guardias,
        'labels': json.dumps(labels),
        'entradas_data': json.dumps(entradas_data),
        'salidas_data': json.dumps(salidas_data),
        'tipos_labels': json.dumps(tipos_labels),
        'tipos_data': json.dumps(tipos_data),
        'estatus_labels': json.dumps(estatus_labels),
        'estatus_data': json.dumps(estatus_data),
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'guardia_id': guardia_id or '',
        'turnos_stats': turnos_stats,
        'usuarios_stats': usuarios_stats,
        'media_url': settings.MEDIA_URL,
    }

    return render(request, 'platform_core/admin_dashboard.html', context)


# =========================
# WEB VIEWS - REGISTROS ACCESO
# =========================

@login_required
def registro_acceso_list_view(request):
    registros = RegistroAcceso.objects.select_related(
        'turno', 'guardia', 'vehiculo', 'empleado', 'conductor', 'visitante'
    ).order_by('-fecha_hora')

    tipo_movimiento = request.GET.get('tipo_movimiento')
    tipo_entidad = request.GET.get('tipo_entidad')
    vehiculo_id = request.GET.get('vehiculo')
    visitante_id = request.GET.get('visitante')
    q = request.GET.get('q')

    if tipo_movimiento:
        registros = registros.filter(tipo_movimiento=tipo_movimiento)
    if tipo_entidad:
        registros = registros.filter(tipo_entidad=tipo_entidad)
    if vehiculo_id:
        registros = registros.filter(vehiculo_id=vehiculo_id)
    if visitante_id:
        registros = registros.filter(visitante_id=visitante_id)
    if q:
        registros = registros.filter(
            Q(vehiculo__placa__icontains=q) |
            Q(conductor__nombre_completo__icontains=q) |
            Q(empleado__nombre_completo__icontains=q) |
            Q(visitante__nombre_completo__icontains=q)
        )

    context = {
        'registros': registros[:100],
        'tipo_movimiento': tipo_movimiento or '',
        'tipo_entidad': tipo_entidad or '',
        'vehiculo_id': vehiculo_id or '',
        'visitante_id': visitante_id or '',
        'q': q or '',
    }

    return render(request, 'platform_core/registro_acceso_list.html', context)


# =========================
# WEB VIEWS - VEHÍCULOS
# =========================

@login_required
def vehiculo_list_view(request):
    vehiculos = Vehiculo.objects.all().order_by('placa')

    tipo_entidad = request.GET.get('tipo_entidad')
    categoria = request.GET.get('categoria')
    activo = request.GET.get('activo')
    q = request.GET.get('q')

    if tipo_entidad:
        vehiculos = vehiculos.filter(tipo_entidad=tipo_entidad)

    if categoria:
        vehiculos = vehiculos.filter(categoria=categoria)

    if activo in ['true', 'false']:
        vehiculos = vehiculos.filter(activo=(activo == 'true'))

    if q:
        vehiculos = vehiculos.filter(
            Q(placa__icontains=q) |
            Q(clave_interna__icontains=q) |
            Q(propietario__icontains=q) |
            Q(empresa__icontains=q)
        )

    context = {
        'vehiculos': vehiculos,
        'tipo_entidad': tipo_entidad or '',
        'categoria': categoria or '',
        'activo': activo or '',
        'q': q or '',
    }

    return render(request, 'platform_core/vehiculo_list.html', context)


@login_required
def vehiculos_en_instalacion_view(request):
    tractos = Vehiculo.objects.filter(
        en_instalacion=True,
        tipo_entidad='tracto'
    ).select_related('conductor_actual').order_by('placa')

    vehiculos_empresa = Vehiculo.objects.filter(
        en_instalacion=True,
        tipo_entidad='empleado'
    ).select_related('ultimo_empleado').order_by('placa')

    from .models import RegistroAcceso
    visitantes_pendientes = RegistroAcceso.objects.filter(
        tipo_entidad='visitante',
        tipo_movimiento='entrada'
    ).exclude(
        entradas_salidas_asociadas__isnull=False
    ).select_related('visitante', 'turno')

    context = {
        'tractos': tractos,
        'vehiculos_empresa': vehiculos_empresa,
        'visitantes_pendientes': visitantes_pendientes,
    }

    return render(request, 'platform_core/vehiculos_en_instalacion.html', context)


@login_required
def vehiculo_create_view(request):
    if request.method == 'POST':
        vehiculo = Vehiculo.objects.create(
            clave_interna=request.POST.get('clave_interna') or None,
            placa=request.POST.get('placa', '').upper(),
            tipo_entidad=request.POST.get('tipo_entidad'),
            categoria=request.POST.get('categoria'),
            propietario=request.POST.get('propietario', ''),
            empresa=request.POST.get('empresa', ''),
            marca=request.POST.get('marca', ''),
            modelo=request.POST.get('modelo', ''),
            color=request.POST.get('color', ''),
            numero_economico=request.POST.get('numero_economico', ''),
            activo=True if request.POST.get('activo') == 'on' else False,
            requiere_checklist=True if request.POST.get('requiere_checklist') == 'on' else False,
            observaciones=request.POST.get('observaciones', ''),
        )

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='crear',
            descripcion=f'Se creó vehículo {vehiculo.placa}.',
            entidad_tipo='Vehiculo',
            entidad_id=vehiculo.id,
            turno=None
        )

        messages.success(request, 'Vehículo creado correctamente.')
        return redirect('web-vehiculos')

    return render(request, 'platform_core/vehiculo_form.html', {
        'titulo': 'Nuevo vehículo',
        'vehiculo': None,
    })


@login_required
def vehiculo_update_view(request, pk):
    vehiculo = get_object_or_404(Vehiculo, pk=pk)

    if request.method == 'POST':
        vehiculo.clave_interna = request.POST.get('clave_interna') or None
        vehiculo.placa = request.POST.get('placa', '').upper()
        vehiculo.tipo_entidad = request.POST.get('tipo_entidad')
        vehiculo.categoria = request.POST.get('categoria')
        vehiculo.propietario = request.POST.get('propietario', '')
        vehiculo.empresa = request.POST.get('empresa', '')
        vehiculo.marca = request.POST.get('marca', '')
        vehiculo.modelo = request.POST.get('modelo', '')
        vehiculo.color = request.POST.get('color', '')
        vehiculo.numero_economico = request.POST.get('numero_economico', '')
        vehiculo.activo = True if request.POST.get('activo') == 'on' else False
        vehiculo.requiere_checklist = True if request.POST.get('requiere_checklist') == 'on' else False
        vehiculo.observaciones = request.POST.get('observaciones', '')
        vehiculo.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='editar',
            descripcion=f'Se actualizó vehículo {vehiculo.placa}.',
            entidad_tipo='Vehiculo',
            entidad_id=vehiculo.id,
            turno=None
        )

        messages.success(request, 'Vehículo actualizado correctamente.')
        return redirect('web-vehiculos')

    return render(request, 'platform_core/vehiculo_form.html', {
        'titulo': 'Editar vehículo',
        'vehiculo': vehiculo,
    })


@login_required
def vehiculo_desactivar_view(request, pk):
    vehiculo = get_object_or_404(Vehiculo, pk=pk)
    vehiculo.activo = False
    vehiculo.save()

    registrar_auditoria(
        usuario=request.user,
        modulo='catalogos',
        accion='editar',
        descripcion=f'Se desactivó vehículo {vehiculo.placa}.',
        entidad_tipo='Vehiculo',
        entidad_id=vehiculo.id,
        turno=None
    )

    messages.warning(request, 'Vehículo desactivado correctamente.')
    return redirect('web-vehiculos')


# =========================
# WEB VIEWS - EMPLEADOS
# =========================

@login_required
def empleado_list_view(request):
    empleados = Empleado.objects.all().order_by('nombre_completo')

    q = request.GET.get('q')
    activo = request.GET.get('activo')
    sucursal = request.GET.get('sucursal')

    if q:
        empleados = empleados.filter(
            Q(nombre_completo__icontains=q) |
            Q(numero_empleado__icontains=q)
        )

    if activo in ['true', 'false']:
        empleados = empleados.filter(activo=(activo == 'true'))

    if sucursal:
        empleados = empleados.filter(departamento=sucursal)

    return render(request, 'platform_core/empleado_list.html', {
        'empleados': empleados,
        'q': q or '',
        'activo': activo or '',
        'sucursal': sucursal or '',
    })


@login_required
def empleado_create_view(request):
    if request.method == 'POST':
        empleado = Empleado.objects.create(
            numero_empleado=request.POST.get('numero_empleado', ''),
            nombre_completo=request.POST.get('nombre_completo', ''),
            departamento=request.POST.get('departamento', ''),
            puesto=request.POST.get('puesto', ''),
            activo=True if request.POST.get('activo') == 'on' else False,
        )

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='crear',
            descripcion=f'Se creó empleado {empleado.nombre_completo}.',
            entidad_tipo='Empleado',
            entidad_id=empleado.id,
            turno=None
        )

        messages.success(request, 'Empleado creado correctamente.')
        return redirect('web-empleados')

    return render(request, 'platform_core/empleado_form.html', {
        'titulo': 'Nuevo empleado',
        'empleado': None,
    })


@login_required
def empleado_update_view(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)

    if request.method == 'POST':
        empleado.numero_empleado = request.POST.get('numero_empleado', '')
        empleado.nombre_completo = request.POST.get('nombre_completo', '')
        empleado.departamento = request.POST.get('departamento', '')
        empleado.puesto = request.POST.get('puesto', '')
        empleado.activo = True if request.POST.get('activo') == 'on' else False
        empleado.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='editar',
            descripcion=f'Se actualizó empleado {empleado.nombre_completo}.',
            entidad_tipo='Empleado',
            entidad_id=empleado.id,
            turno=None
        )

        messages.success(request, 'Empleado actualizado correctamente.')
        return redirect('web-empleados')

    return render(request, 'platform_core/empleado_form.html', {
        'titulo': 'Editar empleado',
        'empleado': empleado,
    })


@login_required
def empleado_desactivar_view(request, pk):
    empleado = get_object_or_404(Empleado, pk=pk)
    empleado.activo = False
    empleado.save()

    registrar_auditoria(
        usuario=request.user,
        modulo='catalogos',
        accion='editar',
        descripcion=f'Se desactivó empleado {empleado.nombre_completo}.',
        entidad_tipo='Empleado',
        entidad_id=empleado.id,
        turno=None
    )

    messages.warning(request, 'Empleado desactivado correctamente.')
    return redirect('web-empleados')


# =========================
# WEB VIEWS - ASIGNACIONES
# =========================

@login_required
def asignacion_list_view(request):
    asignaciones = AsignacionVehiculoEmpleado.objects.select_related(
        'vehiculo',
        'empleado'
    ).all().order_by('-fecha_asignacion')

    q = request.GET.get('q')
    activa = request.GET.get('activa')

    if q:
        asignaciones = asignaciones.filter(
            Q(vehiculo__placa__icontains=q) |
            Q(empleado__nombre_completo__icontains=q) |
            Q(empleado__numero_empleado__icontains=q)
        )

    if activa in ['true', 'false']:
        asignaciones = asignaciones.filter(activa=(activa == 'true'))

    return render(request, 'platform_core/asignacion_list.html', {
        'asignaciones': asignaciones,
        'q': q or '',
        'activa': activa or '',
    })


@login_required
def asignacion_create_view(request):
    empleados = Empleado.objects.filter(activo=True).order_by('nombre_completo')
    vehiculos = Vehiculo.objects.filter(activo=True, tipo_entidad='empleado').order_by('placa')

    if request.method == 'POST':
        asignacion = AsignacionVehiculoEmpleado.objects.create(
            vehiculo_id=request.POST.get('vehiculo'),
            empleado_id=request.POST.get('empleado'),
            activa=True if request.POST.get('activa') == 'on' else False,
        )

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='crear',
            descripcion=f'Se creó asignación vehículo-empleado {asignacion.id}.',
            entidad_tipo='AsignacionVehiculoEmpleado',
            entidad_id=asignacion.id,
            turno=None
        )

        messages.success(request, 'Asignación creada correctamente.')
        return redirect('web-asignaciones')

    return render(request, 'platform_core/asignacion_form.html', {
        'titulo': 'Nueva asignación',
        'asignacion': None,
        'empleados': empleados,
        'vehiculos': vehiculos,
    })


@login_required
def asignacion_update_view(request, pk):
    asignacion = get_object_or_404(AsignacionVehiculoEmpleado, pk=pk)
    empleados = Empleado.objects.filter(activo=True).order_by('nombre_completo')
    vehiculos = Vehiculo.objects.filter(activo=True, tipo_entidad='empleado').order_by('placa')

    if request.method == 'POST':
        asignacion.vehiculo_id = request.POST.get('vehiculo')
        asignacion.empleado_id = request.POST.get('empleado')
        asignacion.activa = True if request.POST.get('activa') == 'on' else False
        asignacion.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='editar',
            descripcion=f'Se actualizó asignación vehículo-empleado {asignacion.id}.',
            entidad_tipo='AsignacionVehiculoEmpleado',
            entidad_id=asignacion.id,
            turno=None
        )

        messages.success(request, 'Asignación actualizada correctamente.')
        return redirect('web-asignaciones')

    return render(request, 'platform_core/asignacion_form.html', {
        'titulo': 'Editar asignación',
        'asignacion': asignacion,
        'empleados': empleados,
        'vehiculos': vehiculos,
    })


@login_required
def asignacion_desactivar_view(request, pk):
    asignacion = get_object_or_404(AsignacionVehiculoEmpleado, pk=pk)
    asignacion.activa = False
    asignacion.save()

    registrar_auditoria(
        usuario=request.user,
        modulo='catalogos',
        accion='editar',
        descripcion=f'Se desactivó asignación vehículo-empleado {asignacion.id}.',
        entidad_tipo='AsignacionVehiculoEmpleado',
        entidad_id=asignacion.id,
        turno=None
    )

    messages.warning(request, 'Asignación desactivada correctamente.')
    return redirect('web-asignaciones')


# =========================
# WEB VIEWS - CONDUCTORES
# =========================

@login_required
def conductor_list_view(request):
    conductores = Conductor.objects.all().order_by('nombre_completo')

    q = request.GET.get('q')
    activo = request.GET.get('activo')

    if q:
        conductores = conductores.filter(
            Q(nombre_completo__icontains=q) |
            Q(licencia__icontains=q) |
            Q(telefono__icontains=q) |
            Q(empresa__icontains=q)
        )

    if activo in ['true', 'false']:
        conductores = conductores.filter(activo=(activo == 'true'))

    return render(request, 'platform_core/conductor_list.html', {
        'conductores': conductores,
        'q': q or '',
        'activo': activo or '',
    })


@login_required
def conductor_create_view(request):
    if request.method == 'POST':
        conductor = Conductor.objects.create(
            nombre_completo=request.POST.get('nombre_completo', ''),
            licencia=request.POST.get('licencia', ''),
            telefono=request.POST.get('telefono', ''),
            empresa=request.POST.get('empresa', ''),
            activo=True if request.POST.get('activo') == 'on' else False,
        )

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='crear',
            descripcion=f'Se creó conductor {conductor.nombre_completo}.',
            entidad_tipo='Conductor',
            entidad_id=conductor.id,
            turno=None
        )

        messages.success(request, 'Conductor creado correctamente.')
        return redirect('web-conductores')

    return render(request, 'platform_core/conductor_form.html', {
        'titulo': 'Nuevo conductor',
        'conductor': None,
    })


@login_required
def conductor_update_view(request, pk):
    conductor = get_object_or_404(Conductor, pk=pk)

    if request.method == 'POST':
        conductor.nombre_completo = request.POST.get('nombre_completo', '')
        conductor.licencia = request.POST.get('licencia', '')
        conductor.telefono = request.POST.get('telefono', '')
        conductor.empresa = request.POST.get('empresa', '')
        conductor.activo = True if request.POST.get('activo') == 'on' else False
        conductor.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='catalogos',
            accion='editar',
            descripcion=f'Se actualizó conductor {conductor.nombre_completo}.',
            entidad_tipo='Conductor',
            entidad_id=conductor.id,
            turno=None
        )

        messages.success(request, 'Conductor actualizado correctamente.')
        return redirect('web-conductores')

    return render(request, 'platform_core/conductor_form.html', {
        'titulo': 'Editar conductor',
        'conductor': conductor,
    })


@login_required
def conductor_desactivar_view(request, pk):
    conductor = get_object_or_404(Conductor, pk=pk)
    conductor.activo = False
    conductor.save()

    registrar_auditoria(
        usuario=request.user,
        modulo='catalogos',
        accion='editar',
        descripcion=f'Se desactivó conductor {conductor.nombre_completo}.',
        entidad_tipo='Conductor',
        entidad_id=conductor.id,
        turno=None
    )

    messages.warning(request, 'Conductor desactivado correctamente.')
    return redirect('web-conductores')


# =========================
# WEB VIEWS - TURNOS
# =========================

@login_required
def turno_list_view(request):
    turnos = Turno.objects.select_related('guardia').all().order_by('-fecha', '-hora_apertura')

    guardia = request.GET.get('guardia')
    tipo_turno = request.GET.get('tipo_turno')
    abierto = request.GET.get('abierto')

    if guardia:
        turnos = turnos.filter(guardia_id=guardia)

    if tipo_turno:
        turnos = turnos.filter(tipo_turno=tipo_turno)

    if abierto in ['true', 'false']:
        turnos = turnos.filter(abierto=(abierto == 'true'))

    return render(request, 'platform_core/turno_list.html', {
        'turnos': turnos,
        'guardias': User.objects.all().order_by('username'),
        'guardia': guardia or '',
        'tipo_turno': tipo_turno or '',
        'abierto': abierto or '',
    })


@login_required
def turno_create_view(request):
    guardias = User.objects.all().order_by('username')

    if request.method == 'POST':
        turno = Turno.objects.create(
            guardia_id=request.POST.get('guardia'),
            tipo_turno=request.POST.get('tipo_turno'),
            fecha=request.POST.get('fecha'),
            hora_apertura=parse_datetime_form(request.POST.get('hora_apertura')),
            observaciones=request.POST.get('observaciones', ''),
            abierto=True if request.POST.get('abierto') == 'on' else False,
        )

        registrar_auditoria(
            usuario=request.user,
            modulo='turnos',
            accion='crear',
            descripcion=f'Se abrió turno {turno.id} de tipo {turno.tipo_turno}.',
            entidad_tipo='Turno',
            entidad_id=turno.id,
            turno=turno
        )

        messages.success(request, 'Turno creado correctamente.')
        return redirect('web-turnos')

    return render(request, 'platform_core/turno_form.html', {
        'titulo': 'Nuevo turno',
        'turno': None,
        'guardias': guardias,
    })


@login_required
def turno_update_view(request, pk):
    turno = get_object_or_404(Turno, pk=pk)
    guardias = User.objects.all().order_by('username')

    if request.method == 'POST':
        turno.guardia_id = request.POST.get('guardia')
        turno.tipo_turno = request.POST.get('tipo_turno')
        turno.fecha = request.POST.get('fecha')
        turno.hora_apertura = parse_datetime_form(request.POST.get('hora_apertura'))
        turno.observaciones = request.POST.get('observaciones', '')
        turno.abierto = True if request.POST.get('abierto') == 'on' else False

        if not turno.abierto and not turno.hora_cierre:
            turno.hora_cierre = timezone.now()

        turno.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='turnos',
            accion='editar',
            descripcion=f'Se actualizó turno {turno.id}.',
            entidad_tipo='Turno',
            entidad_id=turno.id,
            turno=turno
        )

        messages.success(request, 'Turno actualizado correctamente.')
        return redirect('web-turnos')

    return render(request, 'platform_core/turno_form.html', {
        'titulo': 'Editar turno',
        'turno': turno,
        'guardias': guardias,
    })


@login_required
def turno_cerrar_view(request, pk):
    turno = get_object_or_404(Turno, pk=pk)
    turno.abierto = False
    turno.hora_cierre = timezone.now()
    turno.save()

    registrar_auditoria(
        usuario=request.user,
        modulo='turnos',
        accion='cerrar',
        descripcion=f'Se cerró el turno {turno.id}.',
        entidad_tipo='Turno',
        entidad_id=turno.id,
        turno=turno
    )

    messages.warning(request, 'Turno cerrado correctamente.')
    return redirect('web-turnos')


@login_required
def reporte_turno_detalle_view(request, pk):
    from django.contrib.auth import get_user_model
    User = get_user_model()

    turno = get_object_or_404(Turno.objects.select_related('guardia'), pk=pk)

    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, 'No tienes permisos para ver este reporte.')
        return redirect('web-reportes-turnos')

    registros = RegistroAcceso.objects.select_related(
        'vehiculo',
        'empleado',
        'conductor',
        'visitante',
        'guardia',
        'turno',
    ).filter(turno=turno).order_by('fecha_hora')

    checklists = ChecklistTracto.objects.select_related(
        'vehiculo',
        'conductor',
        'guardia',
        'turno',
    ).filter(turno=turno).order_by('fecha_hora')

    export_csv = request.GET.get('export_csv')

    if export_csv:
        import csv
        from django.http import HttpResponse

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="reporte_turno_{turno.id}_{turno.fecha}.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'Fecha', 'Hora', 'Tipo Movimiento', 'Tipo Entidad',
            'Placa', 'Clave Interna', 'Conductor/Empleado/Visitante',
            'Checklist Requerido', 'Checklist Realizado', 'Observaciones'
        ])

        for reg in registros:
            persona = ''
            if reg.conductor:
                persona = reg.conductor.nombre_completo
            elif reg.empleado:
                persona = reg.empleado.nombre_completo
            elif reg.visitante:
                persona = reg.visitante.nombre_completo

            writer.writerow([
                reg.fecha_hora.strftime('%d/%m/%Y'),
                reg.fecha_hora.strftime('%H:%M'),
                reg.tipo_movimiento,
                reg.tipo_entidad,
                reg.vehiculo.placa if reg.vehiculo else '',
                reg.vehiculo.clave_interna if reg.vehiculo else '',
                persona,
                'Sí' if reg.checklist_requerido else 'No',
                'Sí' if reg.checklist_realizado else 'No',
                reg.observaciones or '',
            ])

        return response

    resumen = {
        'total_registros': registros.count(),
        'entradas': registros.filter(tipo_movimiento='entrada').count(),
        'salidas': registros.filter(tipo_movimiento='salida').count(),
        'checklists_realizados': checklists.count(),
        'checklists_aprobados': checklists.filter(estatus_general='aprobado').count(),
        'checklists_rechazados': checklists.filter(estatus_general='rechazado').count(),
        'checklists_condicionados': checklists.filter(estatus_general='condicionado').count(),
    }

    return render(request, 'platform_core/reporte_turno_detalle.html', {
        'turno': turno,
        'registros': registros,
        'checklists': checklists,
        'resumen': resumen,
    })


@login_required
def turno_historial_view(request, pk):
    turno = get_object_or_404(Turno.objects.select_related('guardia'), pk=pk)

    registros = RegistroAcceso.objects.select_related(
        'vehiculo',
        'empleado',
        'conductor',
        'visitante',
        'guardia',
    ).filter(turno=turno).order_by('-fecha_hora')

    checklists = ChecklistTracto.objects.select_related(
        'vehiculo',
        'conductor',
        'guardia',
    ).filter(turno=turno).order_by('-fecha_hora')

    auditorias = AuditLog.objects.select_related(
        'usuario',
        'turno'
    ).filter(turno=turno).order_by('-fecha_hora')

    return render(request, 'platform_core/turno_historial.html', {
        'turno': turno,
        'registros': registros,
        'checklists': checklists,
        'auditorias': auditorias,
    })


# =========================
# WEB - CHECKLIST TRACTO
# =========================

@login_required
def checklist_tracto_list_view(request):
    checklists = ChecklistTracto.objects.select_related(
        'registro_acceso',
        'turno',
        'guardia',
        'vehiculo',
        'conductor'
    ).all().order_by('-fecha_hora', '-id')

    estatus_general = request.GET.get('estatus_general')
    turno = request.GET.get('turno')
    q = request.GET.get('q')

    if estatus_general:
        checklists = checklists.filter(estatus_general=estatus_general)

    if turno:
        checklists = checklists.filter(turno_id=turno)

    if q:
        checklists = checklists.filter(
            Q(vehiculo__placa__icontains=q) |
            Q(conductor__nombre_completo__icontains=q) |
            Q(guardia__username__icontains=q)
        )

    return render(request, 'platform_core/checklist_tracto_list.html', {
        'checklists': checklists,
        'turnos': Turno.objects.all().order_by('-fecha', '-hora_apertura'),
        'estatus_general': estatus_general or '',
        'turno': turno or '',
        'q': q or '',
    })


@login_required
def checklist_tracto_create_view(request):
    registros = RegistroAcceso.objects.select_related(
        'turno',
        'guardia',
        'vehiculo',
        'conductor'
    ).filter(
        tipo_entidad='tracto',
        checklist_requerido=True,
        checklist_realizado=False
    ).order_by('-fecha_hora', '-id')

    catalogo_items = ChecklistTractoItemCatalogo.objects.filter(
        activo=True
    ).order_by('seccion', 'orden', 'id')

    llantas_posiciones = ChecklistTractoLlanta.POSICION_CHOICES

    if request.method == 'POST':
        registro = get_object_or_404(
            RegistroAcceso,
            pk=request.POST.get('registro_acceso')
        )

        checklist = ChecklistTracto.objects.create(
            registro_acceso=registro,
            turno=registro.turno,
            guardia=registro.guardia,
            vehiculo=registro.vehiculo,
            conductor=registro.conductor,
            estatus_general=request.POST.get('estatus_general'),
            observaciones_generales=request.POST.get('observaciones_generales', ''),
            firma_operador_data=request.POST.get('firma_operador_data', ''),
            firma_vigilante_data=request.POST.get('firma_vigilante_data', ''),
        )

        for item in catalogo_items:
            valor = request.POST.get(f'item_{item.id}_valor')
            observacion = request.POST.get(f'item_{item.id}_observacion', '')

            if valor:
                ChecklistTractoResultado.objects.create(
                    checklist=checklist,
                    item=item,
                    valor=valor,
                    observacion=observacion,
                )

        for posicion, _label in llantas_posiciones:
            estado = request.POST.get(f'llanta_{posicion}_estado')
            observacion = request.POST.get(f'llanta_{posicion}_observacion', '')

            if estado:
                ChecklistTractoLlanta.objects.create(
                    checklist=checklist,
                    posicion=posicion,
                    estado=estado,
                    observacion=observacion,
                )

        for img in request.FILES.getlist('evidencias'):
            ChecklistTractoEvidencia.objects.create(
                checklist=checklist,
                imagen=img,
                descripcion=''
            )

        registro.checklist_realizado = True
        registro.save(update_fields=['checklist_realizado'])

        registrar_auditoria(
            usuario=request.user,
            modulo='checklist_tracto',
            accion='crear',
            descripcion=f'Se creó checklist web de tracto {checklist.id}.',
            entidad_tipo='ChecklistTracto',
            entidad_id=checklist.id,
            turno=checklist.turno
        )

        messages.success(request, 'Checklist de tracto creado correctamente.')
        return redirect('web-checklists-tracto')

    secciones = {}
    for item in catalogo_items:
        secciones.setdefault(item.get_seccion_display(), []).append(item)

    return render(request, 'platform_core/checklist_tracto_form.html', {
        'titulo': 'Nuevo checklist de tractocamión',
        'registros': registros,
        'secciones': secciones,
        'llantas_posiciones': llantas_posiciones,
    })


@login_required
def checklist_tracto_detail_view(request, pk):
    checklist = get_object_or_404(
        ChecklistTracto.objects.select_related(
            'registro_acceso',
            'turno',
            'guardia',
            'vehiculo',
            'conductor'
        ).prefetch_related(
            'resultados__item',
            'llantas',
            'evidencias'
        ),
        pk=pk
    )

    return render(request, 'platform_core/checklist_tracto_detail.html', {
        'checklist': checklist,
    })


# =========================
# WEB - GUARDIAS / PERFILES
# =========================

@login_required
def guardia_list_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('web-dashboard')

    usuarios = User.objects.select_related('profile').filter(
        profile__role='guardia'
    ).order_by('username')

    q = request.GET.get('q')

    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) |
            Q(email__icontains=q) |
            Q(profile__full_name__icontains=q)
        )

    return render(request, 'platform_core/guardia_list.html', {
        'usuarios': usuarios,
        'q': q or '',
        'role': 'guardia',
    })


@login_required
def admin_list_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('web-dashboard')

    usuarios = User.objects.select_related('profile').filter(
        profile__role='admin'
    ).order_by('username')

    q = request.GET.get('q')

    if q:
        usuarios = usuarios.filter(
            Q(username__icontains=q) |
            Q(email__icontains=q) |
            Q(profile__full_name__icontains=q)
        )

    return render(request, 'platform_core/admin_list.html', {
        'usuarios': usuarios,
        'q': q or '',
    })


@login_required
def guardia_create_view(request):
    from django.contrib.auth.models import User
    from accounts.models import UserProfile

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        email = request.POST.get('email', '').strip()
        full_name = request.POST.get('full_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        role = request.POST.get('role', 'guardia')

        if not username:
            messages.error(request, 'El nombre de usuario es obligatorio.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Nuevo guardia',
                'is_create': True,
            })

        if not full_name:
            messages.error(request, 'El nombre completo es obligatorio.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Nuevo guardia',
                'is_create': True,
            })

        if User.objects.filter(username=username).exists():
            messages.error(request, 'Este nombre de usuario ya existe.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Nuevo guardia',
                'is_create': True,
            })

        if not password:
            messages.error(request, 'La contraseña es obligatoria.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Nuevo guardia',
                'is_create': True,
            })

        if len(password) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Nuevo guardia',
                'is_create': True,
            })

        if password != password_confirm:
            messages.error(request, 'Las contraseñas no coinciden.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Nuevo guardia',
                'is_create': True,
            })

        usuario = User.objects.create_user(
            username=username,
            password=password,
            email=email
        )

        profile = UserProfile.objects.create(
            user=usuario,
            role=role,
            full_name=full_name,
            phone=phone
        )

        if request.FILES.get('photo'):
            profile.photo = request.FILES.get('photo')
            profile.save()

        registrar_auditoria(
            usuario=request.user,
            modulo='perfil',
            accion='crear',
            descripcion=f'Se creó usuario {username} con rol {role}.',
            entidad_tipo='UserProfile',
            entidad_id=profile.id,
            turno=None
        )

        messages.success(request, f'Usuario {username} creado correctamente.')
        return redirect('web-guardias')

    return render(request, 'platform_core/guardia_form.html', {
        'titulo': 'Nuevo guardia',
        'is_create': True,
    })


@login_required
def guardia_update_view(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    profile = obtener_perfil_usuario(usuario)

    if profile is None:
        messages.error(request, 'Este usuario no tiene perfil asociado.')
        return redirect('web-guardias')

    if request.method == 'POST':
        full_name = request.POST.get('full_name', '').strip()
        email = request.POST.get('email', '').strip()
        phone = request.POST.get('phone', '').strip()
        role = request.POST.get('role', profile.role)
        password_nueva = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')

        if not full_name:
            messages.error(request, 'El nombre completo es obligatorio.')
            return render(request, 'platform_core/guardia_form.html', {
                'titulo': 'Editar guardia',
                'usuario': usuario,
                'profile': profile,
            })

        if password_nueva:
            if len(password_nueva) < 8:
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres.')
                return render(request, 'platform_core/guardia_form.html', {
                    'titulo': 'Editar guardia',
                    'usuario': usuario,
                    'profile': profile,
                })
            if password_nueva != password_confirm:
                messages.error(request, 'Las contraseñas no coinciden.')
                return render(request, 'platform_core/guardia_form.html', {
                    'titulo': 'Editar guardia',
                    'usuario': usuario,
                    'profile': profile,
                })

        usuario.email = email
        usuario.save()

        profile.full_name = full_name
        profile.phone = phone
        profile.role = role

        if request.FILES.get('photo'):
            profile.photo = request.FILES.get('photo')

        profile.save()

        if password_nueva:
            usuario.set_password(password_nueva)
            usuario.save()
            if request.user.id == usuario.id:
                messages.warning(request, 'Contraseña actualizada. Debes iniciar sesión de nuevo.')

        registrar_auditoria(
            usuario=request.user,
            modulo='perfil',
            accion='editar',
            descripcion=f'Se actualizó perfil del usuario {usuario.username}.',
            entidad_tipo='UserProfile',
            entidad_id=profile.id,
            turno=None
        )

        messages.success(request, 'Perfil de guardia actualizado correctamente.')
        return redirect('web-guardias')

    return render(request, 'platform_core/guardia_form.html', {
        'titulo': 'Editar guardia',
        'usuario': usuario,
        'profile': profile,
    })


@login_required
def guardia_delete_view(request, pk):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos.')
        return redirect('web-dashboard')

    usuario = get_object_or_404(User, pk=pk)
    profile = obtener_perfil_usuario(usuario)

    if profile is None or profile.role != 'guardia':
        messages.error(request, 'No puedes eliminar este usuario.')
        return redirect('web-guardias')

    if request.user.id == usuario.id:
        messages.error(request, 'No puedes eliminar tu propio usuario.')
        return redirect('web-guardias')

    username = usuario.username
    profile.delete()
    usuario.delete()

    registrar_auditoria(
        usuario=request.user,
        modulo='perfil',
        accion='eliminar',
        descripcion=f'Se eliminó usuario {username}.',
        entidad_tipo='UserProfile',
        entidad_id=None,
        turno=None
    )

    messages.success(request, f'Usuario {username} eliminado correctamente.')
    return redirect('web-guardias')


# =========================
# WEB - EVIDENCIAS
# =========================

@login_required
def evidencia_list_view(request):
    registros = RegistroAcceso.objects.select_related(
        'turno',
        'vehiculo',
        'empleado',
        'conductor',
        'visitante'
    ).exclude(
        evidencia_fotografica=''
    ).exclude(
        evidencia_fotografica__isnull=True
    ).order_by('-fecha_hora', '-id')

    checklists = ChecklistTracto.objects.prefetch_related(
        'evidencias'
    ).select_related(
        'vehiculo',
        'conductor',
        'guardia',
        'turno'
    ).order_by('-fecha_hora', '-id')

    return render(request, 'platform_core/evidencia_list.html', {
        'registros': registros,
        'checklists': checklists,
    })


# =========================
# WEB - AUDITORÍA
# =========================

@login_required
def audit_list_view(request):
    auditorias = AuditLog.objects.select_related(
        'usuario',
        'turno'
    ).all().order_by('-fecha_hora', '-id')

    modulo = request.GET.get('modulo')
    accion = request.GET.get('accion')
    turno = request.GET.get('turno')
    q = request.GET.get('q')

    if modulo:
        auditorias = auditorias.filter(modulo=modulo)

    if accion:
        auditorias = auditorias.filter(accion=accion)

    if turno:
        auditorias = auditorias.filter(turno_id=turno)

    if q:
        auditorias = auditorias.filter(
            Q(usuario_username_icontains=q) |
            Q(descripcion__icontains=q) |
            Q(entidad_tipo__icontains=q)
        )

    return render(request, 'platform_core/audit_list.html', {
        'auditorias': auditorias,
        'turnos': Turno.objects.all().order_by('-fecha', '-hora_apertura'),
        'modulo': modulo or '',
        'accion': accion or '',
        'turno': turno or '',
        'q': q or '',
    })


# =========================
# WEB - REPORTES
# =========================

@login_required
def reporte_turnos_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos para acceder a reportes.')
        return redirect('web-dashboard')

    from django.contrib.auth import get_user_model
    User = get_user_model()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    guardia_filtro = request.GET.get('guardia')
    estado_filtro = request.GET.get('estado')

    queryset = Turno.objects.select_related('guardia').all().order_by('-fecha', '-hora_apertura')

    if fecha_inicio:
        queryset = queryset.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha__lte=fecha_fin)
    if guardia_filtro:
        queryset = queryset.filter(guardia_id=guardia_filtro)
    if estado_filtro == 'abierto':
        queryset = queryset.filter(abierto=True)
    elif estado_filtro == 'cerrado':
        queryset = queryset.filter(abierto=False)

    resumen = {
        'total': queryset.count(),
        'abiertos': queryset.filter(abierto=True).count(),
        'cerrados': queryset.filter(abierto=False).count(),
        'total_checklists': ChecklistTracto.objects.filter(turno__in=queryset).count(),
    }

    turnos_tipo = queryset.values('tipo_turno').annotate(total=Count('id'))
    turnos_tipo_labels = [t['tipo_turno'] for t in turnos_tipo]
    turnos_tipo_data = [t['total'] for t in turnos_tipo]

    turnos_dia = queryset.values('fecha').annotate(total=Count('id')).order_by('fecha')
    turnos_dia_labels = [t['fecha'].strftime('%d/%m') for t in turnos_dia]
    turnos_dia_data = [t['total'] for t in turnos_dia]

    guardias = User.objects.all().order_by('username')

    context = {
        'turnos': queryset,
        'resumen': resumen,
        'turnos_tipo_labels': json.dumps(turnos_tipo_labels),
        'turnos_tipo_data': json.dumps(turnos_tipo_data),
        'turnos_dia_labels': json.dumps(turnos_dia_labels),
        'turnos_dia_data': json.dumps(turnos_dia_data),
        'guardias': guardias,
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'guardia_filtro': guardia_filtro or '',
        'estado_filtro': estado_filtro or '',
    }

    return render(request, 'platform_core/reporte_turnos.html', context)


@login_required
def reporte_registros_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos para acceder a reportes.')
        return redirect('web-dashboard')

    from django.conf import settings

    User = get_user_model()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    guardia_filtro = request.GET.get('guardia')
    tipo_movimiento = request.GET.get('tipo_movimiento')
    tipo_entidad = request.GET.get('tipo_entidad')

    queryset = RegistroAcceso.objects.select_related(
        'turno', 'guardia', 'vehiculo', 'empleado', 'conductor', 'visitante'
    ).all().order_by('-fecha_hora')

    if fecha_inicio:
        queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)
    if guardia_filtro:
        queryset = queryset.filter(guardia_id=guardia_filtro)
    if tipo_movimiento:
        queryset = queryset.filter(tipo_movimiento=tipo_movimiento)
    if tipo_entidad:
        queryset = queryset.filter(tipo_entidad=tipo_entidad)

    resumen = {
        'total': queryset.count(),
        'entradas': queryset.filter(tipo_movimiento='entrada').count(),
        'salidas': queryset.filter(tipo_movimiento='salida').count(),
        'checklists': queryset.filter(checklist_realizado=True).count(),
    }

    movimientos_hora = queryset.extra(
        {'hora': "date_trunc('hour', fecha_hora)"}
    ).values('hora').annotate(total=Count('id')).order_by('hora')
    movimientos_labels = [m['hora'].strftime('%H:00') for m in movimientos_hora]
    movimientos_data = [m['total'] for m in movimientos_hora]

    tipo_data = queryset.values('tipo_entidad').annotate(total=Count('id'))
    tipo_labels = [t['tipo_entidad'] for t in tipo_data]
    tipo_data_values = [t['total'] for t in tipo_data]

    guardias = User.objects.all().order_by('username')

    context = {
        'registros': queryset[:100],
        'resumen': resumen,
        'movimientos_labels': json.dumps(movimientos_labels),
        'movimientos_data': json.dumps(movimientos_data),
        'tipo_labels': json.dumps(tipo_labels),
        'tipo_data': json.dumps(tipo_data_values),
        'guardias': guardias,
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'guardia_filtro': guardia_filtro or '',
        'tipo_movimiento_filtro': tipo_movimiento or '',
        'tipo_entidad_filtro': tipo_entidad or '',
        'media_url': settings.MEDIA_URL,
    }

    return render(request, 'platform_core/reporte_registros.html', context)


@login_required
def reporte_checklists_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos para acceder a reportes.')
        return redirect('web-dashboard')

    from django.contrib.auth import get_user_model
    User = get_user_model()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    guardia_filtro = request.GET.get('guardia')
    estatus_filtro = request.GET.get('estatus')

    queryset = ChecklistTracto.objects.select_related(
        'vehiculo', 'conductor', 'guardia', 'turno'
    ).all().order_by('-fecha_hora')

    if fecha_inicio:
        queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)
    if guardia_filtro:
        queryset = queryset.filter(guardia_id=guardia_filtro)
    if estatus_filtro:
        queryset = queryset.filter(estatus_general=estatus_filtro)

    resumen = {
        'total': queryset.count(),
        'aprobados': queryset.filter(estatus_general='aprobado').count(),
        'rechazados': queryset.filter(estatus_general='rechazado').count(),
        'condicionados': queryset.filter(estatus_general='condicionado').count(),
    }

    estatus_data = queryset.values('estatus_general').annotate(total=Count('id'))
    estatus_labels = [e['estatus_general'] for e in estatus_data]
    estatus_values = [e['total'] for e in estatus_data]

    checklists_dia = queryset.values('fecha_hora__date').annotate(total=Count('id')).order_by('fecha_hora__date')
    checklists_labels = [c['fecha_hora__date'].strftime('%d/%m') for c in checklists_dia]
    checklists_data = [c['total'] for c in checklists_dia]

    guardias = User.objects.all().order_by('username')

    context = {
        'checklists': queryset[:50],
        'resumen': resumen,
        'estatus_labels': json.dumps(estatus_labels),
        'estatus_values': json.dumps(estatus_values),
        'checklists_labels': json.dumps(checklists_labels),
        'checklists_data': json.dumps(checklists_data),
        'guardias': guardias,
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'guardia_filtro': guardia_filtro or '',
        'estatus_filtro': estatus_filtro or '',
    }

    return render(request, 'platform_core/reporte_checklists.html', context)


@login_required
def reporte_auditoria_view(request):
    if not es_admin(request.user):
        messages.error(request, 'No tienes permisos para acceder a reportes.')
        return redirect('web-dashboard')

    from django.contrib.auth import get_user_model
    User = get_user_model()

    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    modulo_filtro = request.GET.get('modulo')
    accion_filtro = request.GET.get('accion')
    usuario_filtro = request.GET.get('usuario')
    q = request.GET.get('q')

    queryset = AuditLog.objects.select_related('usuario', 'turno').all().order_by('-fecha_hora')

    if fecha_inicio:
        queryset = queryset.filter(fecha_hora__date__gte=fecha_inicio)
    if fecha_fin:
        queryset = queryset.filter(fecha_hora__date__lte=fecha_fin)
    if modulo_filtro:
        queryset = queryset.filter(modulo=modulo_filtro)
    if accion_filtro:
        queryset = queryset.filter(accion=accion_filtro)
    if usuario_filtro:
        queryset = queryset.filter(usuario_id=usuario_filtro)
    if q:
        queryset = queryset.filter(
            Q(descripcion__icontains=q) |
            Q(entidad_tipo__icontains=q)
        )

    resumen = {
        'total': queryset.count(),
        'creaciones': queryset.filter(accion='crear').count(),
        'ediciones': queryset.filter(accion='editar').count(),
        'consultas': queryset.filter(accion='consultar').count(),
    }

    acciones_data = queryset.values('accion').annotate(total=Count('id'))
    acciones_labels = [a['accion'] for a in acciones_data]
    acciones_values = [a['total'] for a in acciones_data]

    modulos_data = queryset.values('modulo').annotate(total=Count('id'))
    modulos_labels = [m['modulo'] for m in modulos_data]
    modulos_values = [m['total'] for m in modulos_data]

    usuarios = User.objects.filter(auditorias__isnull=False).distinct().order_by('username')

    context = {
        'auditorias': queryset[:100],
        'resumen': resumen,
        'acciones_labels': json.dumps(acciones_labels),
        'acciones_values': json.dumps(acciones_values),
        'modulos_labels': json.dumps(modulos_labels),
        'modulos_values': json.dumps(modulos_values),
        'usuarios': usuarios,
        'fecha_inicio': fecha_inicio or '',
        'fecha_fin': fecha_fin or '',
        'modulo_filtro': modulo_filtro or '',
        'accion_filtro': accion_filtro or '',
        'usuario_filtro': usuario_filtro or '',
        'q': q or '',
    }

    return render(request, 'platform_core/reporte_auditoria.html', context)


# =========================
# API: ASIGNACIONES CONDUCTOR-VEHICULO
# =========================

class AsignacionConductorVehiculoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    serializer_class = AsignacionConductorVehiculoSerializer

    def get_queryset(self):
        activa = self.request.query_params.get('activa')
        queryset = AsignacionConductorVehiculo.objects.select_related('conductor', 'vehiculo')
        if activa is not None:
            queryset = queryset.filter(activa=activa.lower() == 'true')
        return queryset.order_by('-fecha_asignacion')


class AsignacionConductorVehiculoCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def post(self, request):
        from platform_core.models import AsignacionConductorVehiculo, BitacoraCambios
        from django.db import transaction

        conductor_id = request.data.get('conductor_id')
        vehiculo_id = request.data.get('vehiculo_id')
        observaciones = request.data.get('observaciones', '')

        if not conductor_id or not vehiculo_id:
            return Response(
                {'error': 'Conductor y vehículo son requeridos.'},
                status=400
            )

        try:
            with transaction.atomic():
                # Verificar si el vehículo ya tiene una asignación activa
                asignacion_activa_vehiculo = AsignacionConductorVehiculo.objects.filter(
                    vehiculo_id=vehiculo_id,
                    activa=True
                ).first()

                if asignacion_activa_vehiculo:
                    return Response(
                        {
                            'error': 'Este vehículo ya tiene una asignación activa.',
                            'asignacion_actual': {
                                'id': asignacion_activa_vehiculo.id,
                                'conductor': asignacion_activa_vehiculo.conductor.nombre_completo
                            }
                        },
                        status=400
                    )

                # Desasignar cualquier asignación activa anterior del conductor
                AsignacionConductorVehiculo.objects.filter(
                    conductor_id=conductor_id,
                    activa=True
                ).update(activa=False, fecha_desasignacion=timezone.now())

                # Crear nueva asignación
                asignacion = AsignacionConductorVehiculo.objects.create(
                    conductor_id=conductor_id,
                    vehiculo_id=vehiculo_id,
                    activa=True,
                    observaciones=observaciones
                )

                # Registrar en bitácora
                BitacoraCambios.registrar(
                    tabla='asignacion_conductor_vehiculo',
                    registro_id=asignacion.id,
                    accion='ASIGNAR',
                    usuario=request.user,
                    datos_nuevos={
                        'conductor_id': conductor_id,
                        'vehiculo_id': vehiculo_id,
                        'observaciones': observaciones
                    }
                )

                serializer = AsignacionConductorVehiculoSerializer(asignacion)
                return Response(serializer.data, status=201)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=500
            )


class AsignacionConductorVehiculoDesasignarAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def post(self, request):
        from platform_core.models import AsignacionConductorVehiculo, BitacoraCambios

        asignacion_id = request.data.get('asignacion_id')

        if not asignacion_id:
            return Response(
                {'error': 'ID de asignación es requerido.'},
                status=400
            )

        try:
            asignacion = AsignacionConductorVehiculo.objects.get(id=asignacion_id)
        except AsignacionConductorVehiculo.DoesNotExist:
            return Response(
                {'error': 'Asignación no encontrada.'},
                status=404
            )

        datos_anteriores = {
            'conductor_id': asignacion.conductor_id,
            'vehiculo_id': asignacion.vehiculo_id,
            'activa': asignacion.activa
        }

        asignacion.activa = False
        asignacion.fecha_desasignacion = timezone.now()
        asignacion.save()

        BitacoraCambios.registrar(
            tabla='asignacion_conductor_vehiculo',
            registro_id=asignacion.id,
            accion='DESASIGNAR',
            usuario=request.user,
            datos_anteriores=datos_anteriores
        )

        return Response({'mensaje': 'Conductor desasignado correctamente.'})


# =========================
# API: ASIGNACIONES EMPLEADO-VEHICULO
# =========================

class AsignacionEmpleadoVehiculoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    serializer_class = AsignacionEmpleadoVehiculoSerializer

    def get_queryset(self):
        activa = self.request.query_params.get('activa')
        queryset = AsignacionEmpleadoVehiculo.objects.select_related('empleado', 'vehiculo')
        if activa is not None:
            queryset = queryset.filter(activa=activa.lower() == 'true')
        return queryset.order_by('-fecha_asignacion')


class AsignacionEmpleadoVehiculoCreateAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def post(self, request):
        from platform_core.models import AsignacionEmpleadoVehiculo, BitacoraCambios
        from django.db import transaction

        empleado_id = request.data.get('empleado_id')
        vehiculo_id = request.data.get('vehiculo_id')
        observaciones = request.data.get('observaciones', '')

        if not empleado_id or not vehiculo_id:
            return Response(
                {'error': 'Empleado y vehículo son requeridos.'},
                status=400
            )

        try:
            with transaction.atomic():
                # Verificar si el empleado ya tiene una asignación activa (regla: 1 empleado = 1 vehículo)
                asignacion_activa_empleado = AsignacionEmpleadoVehiculo.objects.filter(
                    empleado_id=empleado_id,
                    activa=True
                ).first()

                if asignacion_activa_empleado:
                    return Response(
                        {
                            'error': 'Este empleado ya tiene un vehículo asignado.',
                            'asignacion_actual': {
                                'id': asignacion_activa_empleado.id,
                                'vehiculo_placa': asignacion_activa_empleado.vehiculo.placa
                            }
                        },
                        status=400
                    )

                # Verificar si el vehículo ya tiene una asignación activa
                asignacion_activa_vehiculo = AsignacionEmpleadoVehiculo.objects.filter(
                    vehiculo_id=vehiculo_id,
                    activa=True
                ).first()

                if asignacion_activa_vehiculo:
                    return Response(
                        {
                            'error': 'Este vehículo ya tiene una asignación activa.',
                            'asignacion_actual': {
                                'id': asignacion_activa_vehiculo.id,
                                'empleado': asignacion_activa_vehiculo.empleado.nombre_completo
                            }
                        },
                        status=400
                    )

                # Desasignar cualquier asignación activa anterior del empleado
                AsignacionEmpleadoVehiculo.objects.filter(
                    empleado_id=empleado_id,
                    activa=True
                ).update(activa=False, fecha_desasignacion=timezone.now())

                asignacion = AsignacionEmpleadoVehiculo.objects.create(
                    empleado_id=empleado_id,
                    vehiculo_id=vehiculo_id,
                    activa=True,
                    observaciones=observaciones
                )

                BitacoraCambios.registrar(
                    tabla='asignacion_empleado_vehiculo',
                    registro_id=asignacion.id,
                    accion='ASIGNAR',
                    usuario=request.user,
                    datos_nuevos={
                        'empleado_id': empleado_id,
                        'vehiculo_id': vehiculo_id
                    }
                )

                serializer = AsignacionEmpleadoVehiculoSerializer(asignacion)
                return Response(serializer.data, status=201)

        except Exception as e:
            return Response(
                {'error': str(e)},
                status=500
            )


# =========================
# API: HISTORIAL USO VEHÍCULO
# =========================

class HistorialUsoVehiculoListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    serializer_class = HistorialUsoVehiculoSerializer

    def get_queryset(self):
        vehiculo_id = self.request.query_params.get('vehiculo_id')
        dentro = self.request.query_params.get('dentro')
        queryset = HistorialUsoVehiculo.objects.select_related('vehiculo', 'conductor', 'empleado', 'turno')

        if vehiculo_id:
            queryset = queryset.filter(vehiculo_id=vehiculo_id)
        if dentro is not None:
            queryset = queryset.filter(dentro_instalacion=dentro.lower() == 'true')

        return queryset.order_by('-fecha_hora')


class VehiculosDentroAPIView(APIView):
    permission_classes = [IsAuthenticated, IsAdminRole]

    def get(self, request):
        vehiculos = HistorialUsoVehiculo.objects.filter(
            dentro_instalacion=True
        ).select_related('vehiculo', 'conductor', 'empleado', 'turno')

        data = []
        for h in vehiculos:
            data.append({
                'historial_id': h.id,
                'vehiculo_id': h.vehiculo_id,
                'vehiculo_placa': h.vehiculo.placa,
                'vehiculo_marca': h.vehiculo.marca,
                'conductor_id': h.conductor_id,
                'conductor_nombre': h.conductor.nombre_completo if h.conductor else None,
                'empleado_id': h.empleado_id,
                'empleado_nombre': h.empleado.nombre_completo if h.empleado else None,
                'tipo_entidad': h.tipo_entidad,
                'fecha_entrada': h.fecha_hora,
            })

        return Response(data)


# =========================
# API: BITÁCORA CAMBIOS
# =========================

class BitacoraCambiosListAPIView(generics.ListAPIView):
    permission_classes = [IsAuthenticated, IsAdminRole]
    serializer_class = BitacoraCambiosSerializer

    def get_queryset(self):
        tabla = self.request.query_params.get('tabla')
        accion = self.request.query_params.get('accion')
        usuario_id = self.request.query_params.get('usuario_id')

        queryset = BitacoraCambios.objects.select_related('usuario')

        if tabla:
            queryset = queryset.filter(tabla_affectada=tabla)
        if accion:
            queryset = queryset.filter(accion=accion)
        if usuario_id:
            queryset = queryset.filter(usuario_id=usuario_id)

        return queryset.order_by('-fecha_hora')[:100]


# =========================
# WEB: PÁGINA DE ASIGNACIONES
# =========================

@login_required
def admin_asignaciones(request):
    if request.user.profile.role != 'admin':
        return redirect('web-dashboard')

    conductores = Conductor.objects.filter(activo=True).order_by('nombre_completo')
    tractos = Vehiculo.objects.filter(tipo_entidad='tracto', activo=True).order_by('placa')
    empleados = Empleado.objects.filter(activo=True).order_by('nombre_completo')
    vehiculos_empleado = Vehiculo.objects.filter(tipo_entidad='empleado', activo=True).order_by('placa')

    asignaciones_conductor = AsignacionConductorVehiculo.objects.filter(
        activa=True
    ).select_related('conductor', 'vehiculo').order_by('-fecha_asignacion')

    asignaciones_empleado = AsignacionEmpleadoVehiculo.objects.filter(
        activa=True
    ).select_related('empleado', 'vehiculo').order_by('-fecha_asignacion')

    vehiculos_dentro = HistorialUsoVehiculo.objects.filter(
        dentro_instalacion=True
    ).select_related('vehiculo', 'conductor', 'empleado', 'turno').order_by('-fecha_hora')

    bitacora = BitacoraCambios.objects.select_related('usuario').order_by('-fecha_hora')[:50]

    return render(request, 'platform_core/admin_asignaciones.html', {
        'conductores': conductores,
        'tractos': tractos,
        'empleados': empleados,
        'vehiculos_empleado': vehiculos_empleado,
        'asignaciones_conductor': asignaciones_conductor,
        'asignaciones_empleado': asignaciones_empleado,
        'vehiculos_dentro': vehiculos_dentro,
        'bitacora': bitacora,
    })